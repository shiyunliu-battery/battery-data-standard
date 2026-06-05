from __future__ import annotations

import json
import os
import subprocess
import sys
import zipfile
from pathlib import Path

import polars as pl
import pytest

import bds
from battery_data_standard import FileIOError, UnsupportedFeatureError, UnsupportedFormatError
from battery_data_standard.api import batch_convert, list_export_targets, list_supported_formats
from battery_data_standard.audit import audit
from battery_data_standard.profiles import load_profile, profile_column_map
from battery_data_standard.reports import ConversionReport, ValidationReport
from battery_data_standard.schema import BDS_SCHEMA_VERSION


def test_json_profile_maps_common_field_names(tmp_path):
    profile = tmp_path / "profile.json"
    profile.write_text(
        json.dumps(
            {
                "columns": {
                    "test_time": "elapsed_s",
                    "voltage": "Ecell",
                    "current": ["I", "Current"],
                }
            }
        ),
        encoding="utf-8",
    )

    mapped = profile_column_map(load_profile(profile))

    assert mapped["test_time_s"] == ["elapsed_s"]
    assert mapped["voltage_v"] == ["Ecell"]
    assert mapped["current_a"] == ["I", "Current"]


def test_yaml_profile_requires_supported_extension(tmp_path):
    profile = tmp_path / "profile.txt"
    profile.write_text("columns: {}\n", encoding="utf-8")

    with pytest.raises(UnsupportedFormatError, match="Unsupported profile"):
        load_profile(profile)


def test_conversion_report_round_trips_to_json_file(tmp_path):
    report = ConversionReport(
        input_path="raw.csv",
        output_path="out.csv",
        cycler="generic",
        schema_version=BDS_SCHEMA_VERSION,
        rows=2,
        columns=["test_time_s", "voltage_v", "current_a"],
        validation=ValidationReport(
            valid=True,
            schema_version=BDS_SCHEMA_VERSION,
            rows=2,
            columns=["test_time_s", "voltage_v", "current_a"],
        ),
        warnings=["example warning"],
        metadata={"source_backend": "native"},
    )
    path = tmp_path / "report.json"

    report.write_json(path)
    loaded = json.loads(path.read_text(encoding="utf-8"))

    assert loaded["cycler"] == "generic"
    assert loaded["validation"]["valid"] is True
    assert loaded["warnings"] == ["example warning"]


def test_cli_validate_returns_two_for_invalid_file(tmp_path):
    invalid = tmp_path / "invalid.csv"
    pl.DataFrame({"test_time_s": [0.0], "voltage_v": [3.4]}).write_csv(invalid)

    run = subprocess.run(
        [sys.executable, "-m", "battery_data_standard.cli", "validate", str(invalid)],
        text=True,
        capture_output=True,
        check=False,
    )

    assert run.returncode == 4
    assert json.loads(run.stdout)["valid"] is False


def test_missing_input_file_errors_are_clean(tmp_path):
    missing = tmp_path / "missing.csv"

    with pytest.raises(FileIOError, match="Input file does not exist"):
        bds.detect(missing)
    with pytest.raises(FileIOError, match="Input file does not exist"):
        bds.read(missing)


def test_cli_missing_input_file_uses_io_exit_code(tmp_path):
    missing = tmp_path / "missing.csv"

    for command in (
        ["detect", str(missing)],
        ["convert", str(missing), str(tmp_path / "out.csv")],
        ["validate", str(missing)],
    ):
        run = subprocess.run(
            [sys.executable, "-m", "battery_data_standard.cli", *command],
            text=True,
            capture_output=True,
            check=False,
        )

        assert run.returncode == 5
        assert "io error" in run.stderr
        assert "Traceback" not in run.stderr


def test_cli_output_write_failure_uses_io_exit_code(tmp_path):
    raw = tmp_path / "raw.csv"
    raw.write_text("Test Time (s),Voltage (V),Current (A)\n0,3.4,0.1\n", encoding="utf-8")

    run = subprocess.run(
        [
            sys.executable,
            "-m",
            "battery_data_standard.cli",
            "convert",
            str(raw),
            str(tmp_path),
            "--cycler",
            "generic",
        ],
        text=True,
        capture_output=True,
        check=False,
    )

    assert run.returncode == 5
    assert "io error" in run.stderr
    assert "Traceback" not in run.stderr


def test_batch_convert_writes_manifest_and_continues(tmp_path):
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    (input_dir / "ok.csv").write_text("Test Time (s),Voltage (V),Current (A)\n0,3.4,0.1\n", encoding="utf-8")
    (input_dir / "bad.csv").write_text("not,enough\n1,2\n", encoding="utf-8")
    manifest = tmp_path / "manifest.jsonl"

    records = batch_convert(
        input_dir,
        output_dir,
        manifest_path=manifest,
        cycler="generic",
        strict=True,
    )

    assert len(records) == 2
    assert any(record["status"] == "ok" for record in records)
    assert any(record["status"] == "error" for record in records)
    assert manifest.exists()
    assert (output_dir / "ok.bds.csv").exists()


def test_convert_pybamm_target_writes_drive_cycle_table(tmp_path):
    raw = tmp_path / "raw.csv"
    raw.write_text(
        "Test Time (s),Voltage (V),Current (A),Cycle Count\n0,3.4,0.1,1\n5,3.5,0.2,1\n",
        encoding="utf-8",
    )
    output = tmp_path / "drive_cycle.csv"

    report = bds.convert(raw, output, cycler="generic", target="pybamm")

    exported = pl.read_csv(output)
    assert exported.columns == ["time_s", "current_a"]
    assert exported["time_s"].to_list() == [0.0, 5.0]
    assert exported["current_a"].to_list() == [0.1, 0.2]
    assert report.metadata["export_target"] == "pybamm"
    assert report.columns == ["time_s", "current_a"]


def test_convert_pyprobe_target_writes_diagnostic_staging_table(tmp_path):
    raw = tmp_path / "raw.csv"
    raw.write_text(
        "Test Time (s),Voltage (V),Current (A),Cycle Count,Step Index,Step Time (s)\n"
        "0,3.4,0.1,1,1,0\n5,3.5,0.2,1,2,1\n",
        encoding="utf-8",
    )
    output = tmp_path / "diagnostic.parquet"

    report = bds.convert(raw, output, cycler="generic", target="pyprobe", format="parquet")

    exported = pl.read_parquet(output)
    assert exported.columns == [
        "time_s",
        "voltage_v",
        "current_a",
        "cycle_index",
        "step_index",
        "step_time_s",
    ]
    assert report.metadata["export_target"] == "pyprobe"


def test_convert_report_path_auto_writes_json_and_pdf_by_default(tmp_path):
    raw = tmp_path / "raw.csv"
    raw.write_text("Test Time (s),Voltage (V),Current (A)\n0,3.4,0.1\n1,3.5,0.2\n", encoding="utf-8")
    output = tmp_path / "normalized.bds.csv"

    report = bds.convert(raw, output, cycler="generic", report_path="auto")

    json_path = tmp_path / "normalized.bds.report.json"
    pdf_path = tmp_path / "normalized.bds.report.pdf"
    assert output.exists()
    assert json_path.exists()
    assert pdf_path.exists()
    assert not (tmp_path / "normalized.bds.report.html").exists()
    assert not (tmp_path / "normalized.bds.report.xlsx").exists()
    saved = json.loads(json_path.read_text(encoding="utf-8"))
    assert saved["metadata"]["report_outputs"]["json"] == str(json_path)
    assert saved["metadata"]["report_outputs"]["pdf"] == str(pdf_path)
    assert report.metadata["report_outputs"]["pdf"] == str(pdf_path)


def test_convert_report_formats_can_write_blue_html_and_xlsx(tmp_path):
    from openpyxl import load_workbook

    raw = tmp_path / "raw.csv"
    raw.write_text("Test Time (s),Voltage (V),Current (A)\n0,3.4,0.1\n1,3.5,0.2\n", encoding="utf-8")
    output = tmp_path / "normalized.bds.csv"

    bds.convert(raw, output, cycler="generic", report_path="auto", report_formats=("json", "html", "xlsx"))

    html_path = tmp_path / "normalized.bds.report.html"
    xlsx_path = tmp_path / "normalized.bds.report.xlsx"
    pdf_path = tmp_path / "normalized.bds.report.pdf"
    html_text = html_path.read_text(encoding="utf-8")
    workbook = load_workbook(xlsx_path)
    assert "BDS Conversion Report" in html_text
    assert "language result" not in html_text
    assert "#1F4E79" in html_text
    assert workbook["Summary"]["A1"].value == "BDS Conversion Report"
    assert workbook["Summary"]["A1"].fill.fgColor.rgb == "001F4E79"
    assert pdf_path.exists()


def test_convert_repairs_regular_time_sampling_gaps_by_default(tmp_path):
    raw = tmp_path / "gap.csv"
    raw.write_text(
        "Test Time (s),Voltage (V),Current (A)\n0,3.0,0.0\n1,3.1,1.0\n2,3.2,2.0\n4,3.4,4.0\n",
        encoding="utf-8",
    )
    output = tmp_path / "gap.bds.csv"

    report = bds.convert(raw, output, cycler="generic", report_path="auto")
    exported = pl.read_csv(output)

    assert exported["Test Time (s)"].to_list() == [0.0, 1.0, 2.0, 3.0, 4.0]
    assert exported["Voltage (V)"].to_list()[3] == pytest.approx(3.3)
    assert exported["Current (A)"].to_list()[3] == pytest.approx(3.0)
    assert report.metadata["time_sampling"]["status"] == "repaired"
    assert report.metadata["time_sampling"]["missing_points"] == 1
    assert any(issue.code == "missing-sample-timepoints" for issue in report.validation.issues)
    saved = json.loads((tmp_path / "gap.bds.report.json").read_text(encoding="utf-8"))
    assert saved["metadata"]["time_sampling"]["gaps"][0]["missing_times_s"] == [3.0]


def test_convert_time_sampling_warn_policy_reports_without_inserting(tmp_path):
    raw = tmp_path / "gap.csv"
    raw.write_text(
        "Test Time (s),Voltage (V),Current (A)\n0,3.0,0.0\n1,3.1,1.0\n2,3.2,2.0\n4,3.4,4.0\n",
        encoding="utf-8",
    )
    output = tmp_path / "gap.bds.csv"

    report = bds.convert(raw, output, cycler="generic", time_sampling_policy="warn")
    exported = pl.read_csv(output)

    assert exported["Test Time (s)"].to_list() == [0.0, 1.0, 2.0, 4.0]
    assert report.metadata["time_sampling"]["status"] == "gaps-detected"
    assert report.metadata["time_sampling"]["missing_points"] == 1
    assert any("missing sample" in warning for warning in report.warnings)


def test_batch_convert_target_uses_target_suffix(tmp_path):
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    (input_dir / "ok.csv").write_text("Test Time (s),Voltage (V),Current (A)\n0,3.4,0.1\n", encoding="utf-8")

    records = batch_convert(input_dir, output_dir, cycler="generic", target="cellpy")

    assert records[0]["metadata"]["export_target"] == "cellpy"
    assert (output_dir / "ok.cellpy.csv").exists()
    exported = pl.read_csv(output_dir / "ok.cellpy.csv")
    assert exported.columns == ["data_point", "test_time", "current", "voltage"]


def test_batch_convert_rejects_output_file_path(tmp_path):
    input_dir = tmp_path / "input"
    input_dir.mkdir()
    output_file = tmp_path / "not_a_directory"
    output_file.write_text("already here", encoding="utf-8")

    with pytest.raises(FileIOError, match="Output path is not a directory"):
        batch_convert(input_dir, output_file)


def test_cli_batch_writes_outputs_and_manifest(tmp_path):
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    (input_dir / "ok.csv").write_text("Test Time (s),Voltage (V),Current (A)\n0,3.4,0.1\n", encoding="utf-8")
    manifest = tmp_path / "manifest.jsonl"

    run = subprocess.run(
        [
            sys.executable,
            "-m",
            "battery_data_standard.cli",
            "batch",
            str(input_dir),
            str(output_dir),
            "--manifest",
            str(manifest),
            "--cycler",
            "generic",
        ],
        text=True,
        capture_output=True,
        check=True,
    )

    assert json.loads(run.stdout)["files"] == 1
    assert manifest.exists()
    assert (output_dir / "ok.bds.csv").exists()


def test_cli_batch_with_skipped_helper_files_returns_success(tmp_path):
    archive = tmp_path / "records.zip"
    with zipfile.ZipFile(archive, "w") as handle:
        handle.writestr("cell/run.csv", "Test Time (s),Voltage (V),Current (A)\n0,3.4,0.1\n")
        handle.writestr("README.txt", "This archive documents the experiment.\n")
    output_dir = tmp_path / "output"

    run = subprocess.run(
        [
            sys.executable,
            "-m",
            "battery_data_standard.cli",
            "batch",
            str(archive),
            str(output_dir),
            "--cycler",
            "generic",
        ],
        text=True,
        capture_output=True,
        check=True,
    )

    payload = json.loads(run.stdout)
    assert payload["converted"] == 1
    assert payload["skipped"] == 1
    assert payload["errors"] == 0
    assert any(record["record_type"] == "skipped" for record in payload["records"])


def test_cli_batch_with_only_skipped_helper_files_returns_partial(tmp_path):
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    (input_dir / "README.txt").write_text("This directory only contains notes.\n", encoding="utf-8")

    run = subprocess.run(
        [
            sys.executable,
            "-m",
            "battery_data_standard.cli",
            "batch",
            str(input_dir),
            str(output_dir),
        ],
        text=True,
        capture_output=True,
        check=False,
    )

    payload = json.loads(run.stdout)
    assert run.returncode == 6
    assert payload["converted"] == 0
    assert payload["skipped"] == 1
    assert payload["errors"] == 0
    assert payload["records"][0]["record_type"] == "skipped"


def test_audit_scores_converted_unsupported_and_suspicious_files(tmp_path):
    input_dir = tmp_path / "input"
    input_dir.mkdir()
    (input_dir / "ok.csv").write_text(
        "Test Time (s),Voltage (V),Current (A)\n0,3.4,0.1\n1,3.5,0.2\n",
        encoding="utf-8",
    )
    rows = ["Test Time (s),Voltage (V),Current (A),Cycle Count,Step Index"]
    rows.extend(f"{idx if idx != 2 else 1},3.4,0.0,1,{idx}" for idx in range(11))
    (input_dir / "suspicious.csv").write_text("\n".join(rows) + "\n", encoding="utf-8")
    (input_dir / "README.txt").write_text("notes only\n", encoding="utf-8")

    report = audit(input_dir, cycler="generic", current_sign="preserve")

    assert report.files == 2
    assert report.converted == 2
    assert report.unsupported == 0
    suspicious = next(record for record in report.records if record.relative_path == "suspicious.csv")
    assert suspicious.checks["duplicated_timestamps"] == 1
    assert suspicious.checks["suspicious_flat_voltage"]["flag"] is True
    assert any(issue.code == "duplicated-timestamps" for issue in suspicious.issues)
    assert suspicious.quality_score < 100


def test_audit_flags_current_sign_sanity_conflicts(tmp_path):
    raw = tmp_path / "sign_conflict.csv"
    raw.write_text(
        "Test Time (s),Voltage (V),Current (A),Discharging Capacity (Ah)\n"
        "0,3.40,-1.0,0.00\n"
        "1,3.45,-1.0,0.01\n"
        "2,3.50,-1.0,0.02\n"
        "3,3.55,-1.0,0.03\n",
        encoding="utf-8",
    )

    report = audit(raw, cycler="generic", current_sign="charge-positive", current_sign_check="adjacent")
    record = report.records[0]

    assert record.checks["current_sign_confidence"] in {"medium", "high"}
    assert record.checks["current_sign_sanity"]["status"] == "suspicious"
    assert any(issue.code == "current-sign-suspicious" for issue in record.issues)


def test_audit_does_not_overstate_short_current_sign_checks(tmp_path):
    raw = tmp_path / "short.csv"
    raw.write_text(
        "Test Time (s),Voltage (V),Current (A)\n0,3.40,-1.0\n1,3.45,-1.0\n",
        encoding="utf-8",
    )

    report = audit(raw, cycler="generic", current_sign="charge-positive", current_sign_check="adjacent")
    record = report.records[0]

    assert record.checks["current_sign_confidence"] == "inconclusive"
    assert record.checks["current_sign_sanity"]["status"] == "inconclusive"
    assert all(issue.code != "current-sign-suspicious" for issue in record.issues)


def test_audit_can_disable_current_sign_sanity_check(tmp_path):
    raw = tmp_path / "sign_conflict.csv"
    raw.write_text(
        "Test Time (s),Voltage (V),Current (A),Discharging Capacity (Ah)\n"
        "0,3.40,-1.0,0.00\n"
        "1,3.45,-1.0,0.01\n"
        "2,3.50,-1.0,0.02\n"
        "3,3.55,-1.0,0.03\n",
        encoding="utf-8",
    )

    report = audit(raw, cycler="generic", current_sign="charge-positive", current_sign_check="none")
    record = report.records[0]

    assert record.checks["current_sign_confidence"] == "disabled"
    assert record.checks["current_sign_sanity"]["status"] == "disabled"
    assert all(issue.code != "current-sign-suspicious" for issue in record.issues)


def test_audit_flags_non_contiguous_repeated_steps(tmp_path):
    raw = tmp_path / "repeated_steps.csv"
    raw.write_text(
        "Test Time (s),Voltage (V),Current (A),Cycle Count,Step Index\n"
        "0,3.40,0.1,1,1\n"
        "1,3.41,0.1,1,1\n"
        "2,3.42,0.1,1,2\n"
        "3,3.43,0.1,1,2\n"
        "4,3.44,0.1,1,1\n"
        "5,3.45,0.1,1,1\n",
        encoding="utf-8",
    )

    report = audit(raw, cycler="generic", current_sign="preserve")
    record = report.records[0]

    assert record.checks["step_cycle_semantics"]["repeated_step_segments"] == 1
    assert any(issue.code == "repeated-step-segments" for issue in record.issues)


def test_audit_flags_step_transition_discontinuities(tmp_path):
    raw = tmp_path / "step_transition.csv"
    raw.write_text(
        "Test Time (s),Step Time (s),Voltage (V),Current (A),Cycle Count,Step Index\n"
        "0,0,3.40,0.1,1,1\n"
        "1,1,3.41,0.1,1,1\n"
        "2,2,3.42,0.1,1,2\n"
        "3,3,3.43,0.1,1,2\n",
        encoding="utf-8",
    )

    report = audit(raw, cycler="generic", current_sign="preserve")
    record = report.records[0]

    assert record.checks["step_cycle_semantics"]["step_transition_discontinuities"] == 1
    assert any(issue.code == "step-transition-discontinuity" for issue in record.issues)


def test_inferred_test_time_semantics_are_reported(tmp_path):
    raw = tmp_path / "inferred_time.csv"
    raw.write_text(
        "Step Time (s),Voltage (V),Current (A),Step Index\n"
        "0,3.40,0.1,1\n"
        "1,3.41,0.1,1\n"
        "0,3.42,0.1,2\n"
        "1,3.43,0.1,2\n",
        encoding="utf-8",
    )

    _df, conversion_report = bds.read_with_report(
        raw, cycler="generic", current_sign="preserve", strict=False
    )
    audit_report = audit(raw, cycler="generic", current_sign="preserve")
    record = audit_report.records[0]

    assert conversion_report.metadata["semantic_sources"]["test_time_s"]["origin"] == "inferred"
    assert "test_time_s" in record.checks["step_cycle_semantics"]["inferred_fields"]
    assert any(issue.code == "inferred-step-cycle-semantics" for issue in record.issues)


def test_audit_skips_fixture_manifests_and_keeps_optional_completeness_separate():
    fixture_root = Path(__file__).parent / "fixtures"

    report = audit(fixture_root, recursive=True)

    assert report.files >= 1
    assert all(not record.relative_path.endswith("manifest.jsonl") for record in report.records)
    assert "missing-optional-column" not in report.top_issue_codes
    arbin = next(
        record for record in report.records if record.relative_path.replace("\\", "/") == "arbin/basic.csv"
    )
    assert arbin.quality_score >= 90
    assert arbin.quality_grade == "A"
    assert arbin.completeness["required_missing"] == []
    assert "unix_time_s" in arbin.completeness["optional_missing"]


def test_explain_successful_file_reports_mapping_and_next_action():
    source = Path(__file__).parent / "fixtures" / "neware" / "flat.csv"

    report = bds.explain(source)
    payload = report.to_dict()

    assert payload["status"] == "ok"
    assert payload["data_kind"]["kind"] == "timeseries"
    assert payload["selected_adapter"] == "neware"
    assert any(item["source"] == "Current(mA)" for item in payload["unit_transforms"])
    assert any(item["canonical_column"] == "test_time_s" for item in payload["column_mapping"])
    assert "bds convert" in payload["recommended_next_action"]


def test_explain_report_writes_polished_html_json_and_xlsx(tmp_path):
    from openpyxl import load_workbook

    raw = tmp_path / "raw.csv"
    raw.write_text("Test Time (s),Voltage (V),Current (A)\n0,3.4,0.1\n", encoding="utf-8")

    report = bds.explain(raw, cycler="generic")
    outputs = bds.write_explain_reports(report, tmp_path, formats=("json", "html", "xlsx"))

    saved = json.loads(Path(outputs["json"]).read_text(encoding="utf-8"))
    html_text = Path(outputs["html"]).read_text(encoding="utf-8")
    workbook = load_workbook(outputs["xlsx"])

    assert saved["status"] == "ok"
    assert "Report" in html_text
    assert "The source file was identified as" in html_text
    assert "Column Mapping" in html_text
    assert {"Summary", "Column Mapping", "Validation Issues"}.issubset(workbook.sheetnames)
    assert workbook["Summary"]["A1"].value == "BDS Diagnostic Report"
    assert workbook["Column Mapping"]["A1"].value == "Source Column"


def test_explain_failed_file_returns_diagnostic_payload(tmp_path):
    raw = tmp_path / "bad.csv"
    raw.write_text("not,enough\n1,2\n", encoding="utf-8")

    report = bds.explain(raw)
    payload = report.to_dict()

    assert payload["status"] in {"converted-with-issues", "error"}
    assert payload["recommended_next_action"]
    if payload["validation"] is not None:
        assert payload["validation"]["valid"] is False


def test_explain_eis_and_unsupported_files(tmp_path):
    eis = tmp_path / "eis.csv"
    eis.write_text("Frequency_Hz,Zreal_Ohm,Zimag_Ohm\n100,0.1,-0.02\n", encoding="utf-8")
    readme = tmp_path / "README.txt"
    readme.write_text("notes only\n", encoding="utf-8")

    eis_report = bds.explain(eis).to_dict()
    unsupported_report = bds.explain(readme).to_dict()

    assert eis_report["status"] == "eis"
    assert eis_report["data_kind"]["kind"] == "eis"
    assert eis_report["validation"]["valid"] is True
    assert unsupported_report["status"] == "unsupported"
    assert unsupported_report["data_kind"]["kind"] == "unsupported"


def test_cli_explain_json_and_text(tmp_path):
    raw = tmp_path / "raw.csv"
    raw.write_text("Test Time (s),Voltage (V),Current (A)\n0,3.4,0.1\n", encoding="utf-8")
    env = {**os.environ, "PYTHONPATH": str(Path(__file__).parents[1] / "src")}

    json_run = subprocess.run(
        [sys.executable, "-m", "battery_data_standard.cli", "explain", str(raw), "--cycler", "generic"],
        text=True,
        capture_output=True,
        env=env,
        check=True,
    )
    text_run = subprocess.run(
        [
            sys.executable,
            "-m",
            "battery_data_standard.cli",
            "explain",
            str(raw),
            "--cycler",
            "generic",
            "--text",
        ],
        text=True,
        capture_output=True,
        env=env,
        check=True,
    )

    assert json.loads(json_run.stdout)["status"] == "ok"
    assert "BDS explain" in text_run.stdout


def test_cli_explain_writes_user_facing_reports(tmp_path):
    raw = tmp_path / "raw.csv"
    raw.write_text("Test Time (s),Voltage (V),Current (A)\n0,3.4,0.1\n", encoding="utf-8")
    html_path = tmp_path / "explain.html"
    json_path = tmp_path / "explain.json"
    xlsx_path = tmp_path / "explain.xlsx"
    env = {**os.environ, "PYTHONPATH": str(Path(__file__).parents[1] / "src")}

    run = subprocess.run(
        [
            sys.executable,
            "-m",
            "battery_data_standard.cli",
            "explain",
            str(raw),
            "--cycler",
            "generic",
            "--json",
            str(json_path),
            "--html",
            str(html_path),
            "--xlsx",
            str(xlsx_path),
        ],
        text=True,
        capture_output=True,
        env=env,
        check=True,
    )

    assert json.loads(run.stdout)["status"] == "ok"
    assert json.loads(json_path.read_text(encoding="utf-8"))["status"] == "ok"
    assert "BDS Diagnostic Report" in html_path.read_text(encoding="utf-8")
    assert xlsx_path.exists()


def test_detect_kind_uses_schema_aliases_for_common_vendor_headers(tmp_path):
    basytec = tmp_path / "basytec.txt"
    biologic = tmp_path / "bio.mpt"
    novonix = tmp_path / "novonix.csv"
    basytec.write_text("~Time[h]\tU[V]\tI[A]\n0\t3.4\t0.1\n", encoding="utf-8")
    biologic.write_text("time/s\tEwe/V\tI/mA\n0\t3.4\t100\n", encoding="utf-8")
    novonix.write_text("Run Time (h),Potential (V),Current (A)\n0,3.4,0.1\n", encoding="utf-8")

    assert bds.detect_kind(basytec).kind == "timeseries"
    assert bds.detect_kind(biologic).kind == "timeseries"
    assert bds.detect_kind(novonix).kind == "timeseries"


def test_cli_audit_writes_json_and_html(tmp_path):
    input_dir = tmp_path / "input"
    input_dir.mkdir()
    (input_dir / "ok.csv").write_text(
        "Test Time (s),Voltage (V),Current (A)\n0,3.4,0.1\n1,3.5,0.2\n",
        encoding="utf-8",
    )
    json_path = tmp_path / "audit.json"
    html_path = tmp_path / "audit.html"

    run = subprocess.run(
        [
            sys.executable,
            "-m",
            "battery_data_standard.cli",
            "audit",
            str(input_dir),
            "--cycler",
            "generic",
            "--json",
            str(json_path),
            "--html",
            str(html_path),
        ],
        text=True,
        capture_output=True,
        check=True,
    )

    payload = json.loads(run.stdout)
    saved = json.loads(json_path.read_text(encoding="utf-8"))
    assert payload["files"] == 1
    assert saved["records"][0]["status"] == "converted"
    assert "BDS Audit Report" in html_path.read_text(encoding="utf-8")


def test_cli_batch_returns_partial_failure_code(tmp_path):
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    (input_dir / "bad.csv").write_text("not,enough\n1,2\n", encoding="utf-8")

    run = subprocess.run(
        [
            sys.executable,
            "-m",
            "battery_data_standard.cli",
            "batch",
            str(input_dir),
            str(output_dir),
            "--cycler",
            "generic",
        ],
        text=True,
        capture_output=True,
        check=False,
    )

    assert run.returncode == 6
    record = json.loads(run.stdout)["records"][0]
    assert record["status"] == "error"
    assert record["record_type"] == "error"
    assert record["error_type"]


def test_cli_batch_continue_on_error_flag_is_supported(tmp_path):
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    (input_dir / "bad.csv").write_text("not,enough\n1,2\n", encoding="utf-8")

    run = subprocess.run(
        [
            sys.executable,
            "-m",
            "battery_data_standard.cli",
            "batch",
            str(input_dir),
            str(output_dir),
            "--continue-on-error",
            "--cycler",
            "generic",
        ],
        text=True,
        capture_output=True,
        check=False,
    )

    assert run.returncode == 6
    record = json.loads(run.stdout)["records"][0]
    assert record["status"] == "error"
    assert record["record_type"] == "error"
    assert record["error_type"]


def test_batch_records_mpr_missing_optional_backend(tmp_path, monkeypatch):
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    (input_dir / "binary.mpr").write_bytes(b"\x00\x01not a text export")
    manifest = tmp_path / "manifest.jsonl"
    monkeypatch.setitem(sys.modules, "galvani", None)

    records = batch_convert(input_dir, output_dir, manifest_path=manifest)

    assert len(records) == 1
    assert records[0]["status"] == "error"
    assert records[0]["record_type"] == "error"
    assert records[0]["error_type"]
    assert "galvani" in records[0]["error"]
    assert json.loads(manifest.read_text(encoding="utf-8").splitlines()[0])["status"] == "error"


def test_supported_formats_expose_maturity_metadata():
    formats = list_supported_formats()

    assert any(item["cycler"] == "neware" and item["support_tier"] == "fixture-backed" for item in formats)
    biologic = next(item for item in formats if item["cycler"] == "biologic")
    assert ".mpt" in biologic["extensions"]
    assert ".mpr" in biologic["extensions"]
    assert not biologic["unsupported_extensions"]
    assert any(item["cycler"] == "repower" for item in formats)
    assert any(item["cycler"] == "pec" for item in formats)


def test_export_targets_are_discoverable():
    targets = list_export_targets()
    ids = {target["id"] for target in targets}

    assert {
        "bds",
        "bdf",
        "duckdb",
        "polars",
        "cellpy",
        "beep",
        "pybamm",
        "pyprobe",
        "battery-archive",
    }.issubset(ids)


def test_batch_convert_bdf_target_keeps_legacy_suffix(tmp_path):
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    (input_dir / "ok.csv").write_text("Test Time (s),Voltage (V),Current (A)\n0,3.4,0.1\n", encoding="utf-8")

    records = batch_convert(input_dir, output_dir, cycler="generic", target="bdf")

    assert records[0]["metadata"]["export_target"] == "bdf"
    assert records[0]["data_kind"] == "timeseries"
    assert (output_dir / "ok.bdf.csv").exists()
    exported = pl.read_csv(output_dir / "ok.bdf.csv")
    assert {"Test Time / s", "Voltage / V", "Current / A"}.issubset(exported.columns)
    assert "Voltage (V)" not in exported.columns


def test_biologic_mpr_requires_optional_backend(tmp_path, monkeypatch):
    path = tmp_path / "binary.mpr"
    path.write_bytes(b"\x00\x01not a text export")
    monkeypatch.setitem(sys.modules, "galvani", None)

    with pytest.raises(UnsupportedFeatureError, match="galvani"):
        bds.read(path, cycler="biologic")


def test_excel_multiple_plausible_data_sheets_requires_explicit_export(tmp_path):
    pd = pytest.importorskip("pandas")
    path = tmp_path / "multi_sheet.xlsx"
    frame = pd.DataFrame(
        {
            "Test Time (s)": [0, 1],
            "Voltage (V)": [3.4, 3.5],
            "Current (A)": [0.1, 0.2],
        }
    )
    with pd.ExcelWriter(path) as writer:
        frame.to_excel(writer, sheet_name="record_a", index=False)
        frame.to_excel(writer, sheet_name="record_b", index=False)

    with pytest.raises(UnsupportedFormatError, match="multiple plausible data sheets"):
        bds.read(path, cycler="generic")


def test_step_and_cycle_summaries():
    df = pl.DataFrame(
        {
            "test_time_s": [0.0, 10.0, 20.0],
            "voltage_v": [3.4, 3.5, 3.45],
            "current_a": [-0.1, -0.1, 0.2],
            "cycle_index": [1, 1, 1],
            "step_index": [1, 1, 2],
            "charge_capacity_ah": [0.0, 0.01, 0.01],
        }
    )

    steps = bds.summarize_steps(df)
    cycles = bds.summarize_cycles(df)

    assert steps.height == 2
    assert cycles.height == 1
    assert cycles["Duration / s"].to_list() == [20.0]

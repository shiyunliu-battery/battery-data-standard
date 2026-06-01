from __future__ import annotations

import json
import subprocess
import sys
import types
import zipfile

import pytest

import battery_data_standard as bds
from battery_data_standard.archive import DEFAULT_MAX_MEMBER_BYTES, DEFAULT_MAX_TOTAL_BYTES
from battery_data_standard.validation import validate


def test_eis_csv_standardizes_to_eisfit_columns(tmp_path):
    raw = tmp_path / "impedance.csv"
    raw.write_text(
        "Frequency(Hz),R(ohm),X(ohm),SOC\n1000,0.02,-0.001,50\n10,0.05,-0.02,50\n",
        encoding="utf-8",
    )

    kind = bds.detect_kind(raw)
    df = bds.read_eis(raw)

    assert kind.kind == "eis"
    assert df.columns[:5] == ["Frequency_Hz", "Zre_exp_Ohm", "Zim_exp_Ohm", "-Zim_exp_Ohm", "Phase_exp_deg"]
    assert df["Frequency_Hz"].to_list() == [1000.0, 10.0]
    assert df["Zim_exp_Ohm"].to_list() == [-0.001, -0.02]
    assert bds.validate_eis(df).valid


def test_eis_negative_imaginary_column_is_converted_to_complex_imaginary(tmp_path):
    raw = tmp_path / "geis.csv"
    raw.write_text("freq/Hz,Re(Z)/Ohm,-Im(Z)/Ohm\n100,0.1,0.02\n", encoding="utf-8")

    df = bds.read_eis(raw)

    assert df["Zim_exp_Ohm"].to_list() == [-0.02]
    assert df["-Zim_exp_Ohm"].to_list() == [0.02]


def test_batch_archive_routes_bdf_eis_and_unsupported(tmp_path):
    archive = tmp_path / "records.zip"
    with zipfile.ZipFile(archive, "w") as handle:
        handle.writestr("cell/run.csv", "Test Time (s),Voltage (V),Current (A)\n0,3.4,0.1\n")
        handle.writestr("cell/eis.csv", "Frequency_Hz,Zreal_Ohm,Zimag_Ohm\n1000,0.02,-0.001\n")
        handle.writestr("README.txt", "This archive documents the experiment.\n")
    out = tmp_path / "out"
    manifest = tmp_path / "manifest.jsonl"

    records = bds.batch_convert(archive, out, manifest_path=manifest, cycler="generic")

    assert {record["data_kind"] for record in records} == {"timeseries", "eis", "unsupported"}
    assert any(str(record.get("output_path", "")).endswith(".bds.csv") for record in records)
    assert any(str(record.get("output_path", "")).endswith(".eis.csv") for record in records)
    assert any(record.get("archive_member") == "cell/run.csv" for record in records)
    assert all(record["record_type"] == "converted" for record in records if record["status"] == "ok")
    skipped = [record for record in records if record["status"] == "unsupported"]
    assert len(skipped) == 1
    assert skipped[0]["record_type"] == "skipped"
    assert skipped[0]["skip_reason"]
    assert skipped[0]["kind_confidence"] > 0
    assert len(manifest.read_text(encoding="utf-8").splitlines()) == 3


def test_archive_defaults_allow_100_gib_members():
    gib = 1024 * 1024 * 1024

    assert 100 * gib == DEFAULT_MAX_MEMBER_BYTES
    assert 100 * gib == DEFAULT_MAX_TOTAL_BYTES


def test_semicolon_delimited_bms_log_converts(tmp_path):
    raw = tmp_path / "semicolon.csv"
    raw.write_text(
        "Time;s;Test step;Voltage (V);Current (A)\n0;1;rest;3.4;0.1\n1;1;rest;3.5;0.2\n", encoding="utf-8"
    )

    df = bds.read(raw, cycler="generic", current_sign="preserve")

    assert df["test_time_s"].to_list() == [0.0, 1.0]
    assert df["voltage_v"].to_list() == [3.4, 3.5]
    assert validate(df).valid


def test_timestamp_as_header_positional_csv_converts(tmp_path):
    raw = tmp_path / "557_Charge10.csv"
    raw.write_text(
        "\n".join(
            [
                "11/7/2018 5:08:07 PM,3,PAU,00:00:00.000,00:00:00.000,0,0,NN_Char_Charge,3.38427,0.00000,38.06739,0,0,2,",
                "11/7/2018 5:08:08 PM,3,CHA,00:00:01.000,00:00:01.000,0,0,NN_Char_Charge,3.39000,1.50000,38.10000,0,0,2,",
            ]
        ),
        encoding="utf-8",
    )

    df = bds.read(raw, cycler="generic", current_sign="preserve")

    assert df["test_time_s"].to_list() == [0.0, 1.0]
    assert df["voltage_v"].to_list() == pytest.approx([3.38427, 3.39])
    assert df["current_a"].to_list() == pytest.approx([0.0, 1.5])
    assert validate(df).valid


def test_basytec_unit_style_aliases_convert(tmp_path):
    raw = tmp_path / "basytec.txt"
    raw.write_text(
        "~Time[h]\tDataSet\tDateTime\tt-Step[h]\tLine\tCommand\tU[V]\tI[A]\tAh[Ah]\tT1[°C]\tCyc-Count\n"
        "0\t1\t2026-01-01 00:00:00\t0\t1\tRest\t3.4\t0.0\t0\t25\t1\n"
        "0.001\t1\t2026-01-01 00:00:03\t0.001\t1\tCharge\t3.5\t0.2\t0.1\t26\t1\n",
        encoding="utf-8",
    )

    df = bds.read(raw, cycler="basytec", current_sign="preserve")

    assert df["test_time_s"].to_list() == pytest.approx([0.0, 3.6])
    assert df["voltage_v"].to_list() == [3.4, 3.5]
    assert df["current_a"].to_list() == [0.0, 0.2]
    assert validate(df).valid


def test_basytec_tilde_preamble_whitespace_export_converts(tmp_path):
    raw = tmp_path / "basytec_preamble.txt"
    raw.write_text(
        "~Resultfile from Basytec Battery Test System\n"
        "~Start of Test: 14.11.2022 10:40:01\n"
        "~\n"
        "~Time[h] DataSet t-Set[h] Line Command U[V] I[A] T1[°C] Cyc-Count State\n"
        "0 1 0 1 Pause 3.4 0 25 1 1\n"
        "0.001 2 0.001 2 Charge 3.5 0.2 26 1 1\n",
        encoding="utf-8",
    )

    df = bds.read(raw, cycler="basytec", current_sign="preserve")

    assert df["test_time_s"].to_list() == pytest.approx([0.0, 3.6])
    assert df["voltage_v"].to_list() == [3.4, 3.5]
    assert df["current_a"].to_list() == [0.0, 0.2]
    assert "date_time" in df.columns
    assert validate(df).valid


def test_machine_name_columns_convert_units(tmp_path):
    raw = tmp_path / "machine_names.csv"
    raw.write_text(
        "test_time_millisecond,voltage_volt,current_ampere\n0,3.4,0.1\n500,3.5,0.2\n", encoding="utf-8"
    )

    df = bds.read(raw, cycler="generic", current_sign="preserve")

    assert df["test_time_s"].to_list() == [0.0, 0.5]
    assert validate(df).valid


def test_semicolon_with_comma_unit_headers_converts(tmp_path):
    raw = tmp_path / "DrivingAgeing_T5_SOC10-90_TUV_Cell120_03.csv"
    raw.write_text(
        "Time,s;Test step;KAPA_CC,Ah;KAPA_CV,Ah;E,J;I,A;P,W;Q,As;Temp[1],C;U,V\n"
        "0;Rest;0,0;0,0;0,0;0,0;0,0;0,0;25,0;3,4\n"
        "1;Discharge;0,0;0,0;3,5;-0,5;-1,75;0,5;25,1;3,5\n",
        encoding="utf-8",
    )

    df = bds.read(raw, cycler="generic", current_sign="preserve")

    assert df["test_time_s"].to_list() == [0.0, 1.0]
    assert df["voltage_v"].to_list() == [3.4, 3.5]
    assert df["current_a"].to_list() == [0.0, -0.5]
    assert df["power_w"].to_list() == [0.0, -1.75]
    assert validate(df).valid


def test_wonik_stepno_tot_time_columns_convert(tmp_path):
    raw = tmp_path / "05_LHMF_2.csv"
    raw.write_text(
        "DataSequence,Channel,StepNo,Type,StepTime(H:M:S),TotTime(H:M:S),Voltage(V),Current(A),"
        "Capacity(Ah),Power(W),wattHour(Wh),CurCycle,TotCycle,Char. Cap.(Ah),Dischar. Cap.(Ah)\n"
        "1,1,1,Rest,00:00:00,00:00:00,3.4,0.0,0.0,0.0,0.0,1,1,0.0,0.0\n"
        "2,1,2,Discharge,00:00:01,00:00:01,3.5,-0.5,0.1,-1.75,0.175,1,1,0.0,0.1\n",
        encoding="utf-8",
    )

    df = bds.read(raw, cycler="generic", current_sign="preserve")

    assert df["test_time_s"].to_list() == [0.0, 1.0]
    assert df["step_index"].to_list() == [1, 2]
    assert df["cycle_index"].to_list() == [1, 1]
    assert df["discharge_capacity_ah"].to_list() == [0.0, 0.1]
    assert validate(df).valid


def test_neware_channel_voltage_current_columns_convert(tmp_path):
    raw = tmp_path / "HE4C24B02_T45_S000_100_Cycle11.csv"
    raw.write_text(
        "Absolute Time,Testtime,StepID,Steptime,Channel Voltage,Channel Current,Temperature,AuxU\n"
        "2026-01-01 00:00:00,0,1,0,3.4,0.1,25,0\n"
        "2026-01-01 00:00:01,1,1,1,3.5,0.2,25,0\n",
        encoding="utf-8",
    )

    df = bds.read(raw, cycler="generic", current_sign="preserve")

    assert df["test_time_s"].to_list() == [0.0, 1.0]
    assert df["voltage_v"].to_list() == [3.4, 3.5]
    assert df["current_a"].to_list() == [0.1, 0.2]
    assert validate(df).valid


def test_record_number_relative_time_columns_convert(tmp_path):
    raw = tmp_path / "750cycles_58.csv"
    raw.write_text(
        "Record_number,current_A,voltage_V,capacity_Ah,energy_Wh,Relative_Time_h_min_s.ms_,"
        "temperature_C,SOC,Number_of_Cycle\n"
        "1,0.1,3.4,0.0,0.0,00:00:00.000,25,100,1\n"
        "2,0.2,3.5,0.01,0.035,00:00:01.000,25,99,1\n",
        encoding="utf-8",
    )

    df = bds.read(raw, cycler="generic", current_sign="preserve")

    assert df["test_time_s"].to_list() == [0.0, 1.0]
    assert df["record_index"].to_list() == [1, 2]
    assert df["cycle_index"].to_list() == [1, 1]
    assert validate(df).valid


def test_autolab_working_electrode_columns_convert(tmp_path):
    raw = tmp_path / "Charge_top_up.txt"
    raw.write_text(
        "Time (s)\tWE(1).Current (A)\tWE(1).Potential (V)\tWE(1).Charge (C)\n"
        "0\t0.1\t3.4\t0\n"
        "1\t0.2\t3.5\t0.2\n",
        encoding="utf-8",
    )

    df = bds.read(raw, cycler="generic", current_sign="preserve")

    assert df["voltage_v"].to_list() == [3.4, 3.5]
    assert df["current_a"].to_list() == [0.1, 0.2]
    assert validate(df).valid


def test_arbin_underscore_export_aliases_and_sign_convention(tmp_path):
    raw = tmp_path / "Talos_001385_NCR18650_CH33.csv"
    raw.write_text(
        "Data_Point,Test_Time,DateTime,Step_Time,Step_Index,Cycle_Index,Current,Voltage,"
        "Charge_Capacity,Discharge_Capacity,Charge_Energy,Discharge_Energy,Internal_Resistance,"
        "Temperature\n"
        "1,0,2026-01-01 00:00:00,0,1,1,1.0,3.4,0.0,0.0,0.0,0.0,0.01,25\n"
        "2,1,2026-01-01 00:00:01,1,1,1,-0.5,3.5,0.1,0.0,0.3,0.0,0.01,25\n",
        encoding="utf-8",
    )

    detected = bds.detect(raw)
    df = bds.read(raw, cycler="auto")

    assert detected.cycler == "arbin"
    assert df["current_a"].to_list() == [1.0, -0.5]
    assert df["internal_resistance_ohm"].to_list() == [0.01, 0.01]
    assert validate(df).valid


def test_arbin_xlsx_detects_from_sampled_sheet_headers(tmp_path):
    from openpyxl import Workbook

    raw = tmp_path / "M1_Mixed_Aged_R0_T10.xlsx"
    workbook = Workbook()
    info = workbook.active
    info.title = "Global_Info"
    info.append([None, None, None, "TEST REPORT"])
    info.append([None, None, None, "TEST_37_MIX_R0_T10_AGED_DOE2", "Serial Number"])
    info.append(["Channel", "Start DateTime", "Schedule File Name", "Software Version"])
    info.append([1, "2023-02-04 16:20:43", r"C:\ArbinSoftware\DataPro\Result\test", "Mits7"])
    data = workbook.create_sheet("Data")
    data.append(
        [
            "Date_Time",
            "Test_Time(s)",
            "Step_Time(s)",
            "Step_Index",
            "Cycle_Index",
            "Voltage(V)",
            "Current(A)",
            "Charge_Capacity(Ah)",
            "Discharge_Capacity(Ah)",
            "Charge_Energy(Wh)",
            "Discharge_Energy(Wh)",
        ]
    )
    data.append(["2023-02-04 16:20:14.190", 1.0003, 1.0002, 1, 1, 3.5724, 0.0, 0, 0, 0, 0])
    data.append(["2023-02-04 16:20:15.190", 2.0003, 2.0002, 1, 1, 3.5726, -1.0, 0, 0, 0, 0])
    workbook.save(raw)

    detected = bds.detect(raw)
    df = bds.read(raw, cycler="auto", current_sign="preserve", repair_policy="repair")

    assert detected.cycler == "arbin"
    assert df["test_time_s"].to_list() == pytest.approx([0.0, 1.0])
    assert df["current_a"].to_list() == [0.0, -1.0]
    assert validate(df).valid


def test_arbin_xlsx_prefers_channel_sheet_over_rawdata_sheet(tmp_path):
    from openpyxl import Workbook

    raw = tmp_path / "arbin_multisheet.xlsx"
    workbook = Workbook()
    info = workbook.active
    info.title = "Global_Info"
    info.append(["TEST REPORT"])
    rawdata = workbook.create_sheet("RawData_6_1")
    rawdata.append(["Test_Time(s)", "Step_Time(s)", "Cycle_Index", "Step_Index", "Voltage(V)"])
    rawdata.append([1.0, 1.0, 1, 1, 3.4])
    channel = workbook.create_sheet("Channel_6_1")
    channel.append(
        ["Date_Time", "Test_Time(s)", "Step_Time(s)", "Step_Index", "Cycle_Index", "Voltage(V)", "Current(A)"]
    )
    channel.append(["2022-01-01 00:00:00", 1.0, 1.0, 1, 1, 3.4, 0.1])
    channel.append(["2022-01-01 00:00:01", 2.0, 2.0, 1, 1, 3.5, -0.2])
    workbook.save(raw)

    df = bds.read(raw, cycler="auto", current_sign="preserve", repair_policy="repair")

    assert df["test_time_s"].to_list() == [0.0, 1.0]
    assert df["current_a"].to_list() == [0.1, -0.2]
    assert validate(df).valid


def test_arbin_acim_eis_sheet_with_magnitude_phase_converts(tmp_path):
    from openpyxl import Workbook

    raw = tmp_path / "arbin_eis.xlsx"
    workbook = Workbook()
    info = workbook.active
    info.title = "Global_Info"
    info.append(["TEST REPORT"])
    channel = workbook.create_sheet("Channel_6_1")
    channel.append(["Test_Time(s)", "Voltage(V)", "Current(A)"])
    channel.append([0.0, 3.4, 0.0])
    acim = workbook.create_sheet("ACIM_Chan6")
    acim.append(["Device_ID", "Test_ID", "Channel_ID", "Cycle_ID", "Step_ID", "Pt", "Freq", "Zmod", "Zphz"])
    acim.append(["Gamry_1", 1, 6, 1, 2, 0, 1000.0, 0.02, 0.0])
    acim.append(["Gamry_1", 1, 6, 1, 2, 1, 100.0, 0.03, 90.0])
    workbook.save(raw)

    kind = bds.detect_kind(raw)
    df = bds.read_eis(raw)

    assert kind.kind == "eis"
    assert df["Frequency_Hz"].to_list() == [1000.0, 100.0]
    assert df["Zre_exp_Ohm"].to_list()[0] == pytest.approx(0.02)
    assert df["Zim_exp_Ohm"].to_list()[1] == pytest.approx(0.03)
    assert bds.validate_eis(df).valid


def test_maccor_numeric_extension_and_short_headers_convert(tmp_path):
    raw = tmp_path / "xTESLADIAG_000020_CH71.071"
    raw.write_text(
        "Today's Date 07/31/2019  Date of Test:\t07/11/2019\tFilename:\tCH71.071\n"
        "Rec#\tCyc#\tStep\tTest (Sec)\tStep (Sec)\tAmp-hr\tWatt-hr\tAmps\tVolts\tState\t"
        "DPt Time\tACImp/Ohms\tDCIR/Ohms\tTemp 1\n"
        "1\t1\t1\t0\t0\t0\t0\t1.0\t3.4\tC\t07/11/2019 01:00:00 PM\t0.02\t0.03\t25\n"
        "2\t1\t2\t1\t1\t0\t0\t-0.5\t3.5\tD\t07/11/2019 01:00:01 PM\t0.02\t0.03\t25\n",
        encoding="utf-8",
    )

    records = bds.batch_convert(raw, tmp_path / "out")
    df = bds.read(raw, cycler="auto")

    assert records[0]["status"] == "ok"
    assert records[0]["data_kind"] == "timeseries"
    assert records[0]["cycler"] == "maccor"
    assert df["record_index"].to_list() == [1, 2]
    assert df["current_a"].to_list() == [1.0, -0.5]
    assert validate(df).valid


def test_biologic_common_mpt_aliases_convert(tmp_path):
    raw = tmp_path / "bio.mpt"
    raw.write_text(
        "mode\tNs\ttime/s\tEcell/V\t<I>/mA\tQ discharge/mA.h\tQ charge/mA.h\tcycle number\t"
        "P/W\tR/Ohm\tTemperature/°C\n"
        "1\t1\t0\t3.4\t100\t0\t0\t1\t0.34\t0.01\t25\n"
        "1\t1\t1\t3.5\t-50\t0.1\t0\t1\t-0.175\t0.01\t25\n",
        encoding="utf-8",
    )

    df = bds.read(raw, cycler="auto", current_sign="preserve")

    assert df["current_a"].to_list() == [0.1, -0.05]
    assert df["discharge_capacity_ah"].to_list() == [0.0, 0.0001]
    assert df["power_w"].to_list() == [0.34, -0.175]
    assert validate(df).valid


def test_biologic_mpr_uses_galvani_backend(tmp_path, monkeypatch):
    raw = tmp_path / "bio.mpr"
    raw.write_bytes(b"fake mpr bytes")

    class FakeMPRfile:
        def __init__(self, filename):
            self.filename = filename
            self.version = "fixture"
            self.data = [
                {"time/s": 0.0, "Ewe/V": 3.4, "I/mA": 100.0, "cycle number": 1, "Ns": 1},
                {"time/s": 1.0, "Ewe/V": 3.5, "I/mA": -50.0, "cycle number": 1, "Ns": 1},
            ]

    fake_galvani = types.SimpleNamespace(BioLogic=types.SimpleNamespace(MPRfile=FakeMPRfile))
    monkeypatch.setitem(sys.modules, "galvani", fake_galvani)

    df = bds.read(raw, cycler="biologic", current_sign="preserve")

    assert df["test_time_s"].to_list() == [0.0, 1.0]
    assert df["voltage_v"].to_list() == [3.4, 3.5]
    assert df["current_a"].to_list() == [0.1, -0.05]
    assert validate(df).valid


def test_gamry_dta_zcurve_converts_to_eis(tmp_path):
    raw = tmp_path / "gamry_eis.DTA"
    raw.write_text(
        "EXPLAIN\tLABEL\tGamry EIS\n"
        "STARTTIME\tLABEL\t2025-01-01 01:02:03 PM\n"
        "ZCURVE\tTABLE\n"
        "Pt\tTime\tFreq\tZreal\tZimag\tZmod\tZphz\tIdc\tVdc\n"
        "#\ts\tHz\tohm\tohm\tohm\tdeg\tA\tV\n"
        "0\t0\t1000\t0.02\t-0.001\t0.02002\t-2.86\t0\t3.4\n"
        "1\t1\t100\t0.03\t-0.002\t0.03007\t-3.81\t0\t3.4\n"
        "EXPERIMENTABORTED\tFALSE\n",
        encoding="utf-8",
    )

    kind = bds.detect_kind(raw)
    df = bds.read_eis(raw)

    assert kind.kind == "eis"
    assert df["Frequency_Hz"].to_list() == [1000.0, 100.0]
    assert df["Zre_exp_Ohm"].to_list() == [0.02, 0.03]
    assert df["Zim_exp_Ohm"].to_list() == [-0.001, -0.002]
    assert bds.validate_eis(df).valid


def test_repower_csv_aliases_convert(tmp_path):
    raw = tmp_path / "repower.csv"
    raw.write_text(
        "System Time,Relative Time(Sec),Voltage(V),Current(A),Cycle ID,Step ID,Step State,MTV1,"
        "Charge Capacity(Ah),Discharge Capacity(Ah)\n"
        " 2025-01-01 00:00:00,0,3.4,0.1,1,1,Charge,25,0.0,0.0\n"
        " 2025-01-01 00:00:01,1,3.5,0.2,1,2,Discharge,26,0.1,0.0\n",
        encoding="latin-1",
    )

    detected = bds.detect(raw)
    df = bds.read(raw, cycler="auto")

    assert detected.cycler == "repower"
    assert df["test_time_s"].to_list() == [0.0, 1.0]
    assert df["voltage_v"].to_list() == [3.4, 3.5]
    assert df["current_a"].to_list() == [0.1, -0.2]
    assert df["ambient_temperature_deg_c"].to_list() == [25.0, 26.0]
    assert validate(df).valid


def test_pec_csv_aliases_convert(tmp_path):
    raw = tmp_path / "pec.csv"
    raw.write_text(
        "TestRegime Name,Formation\n"
        "Start Time,2025-01-01 00:00:00\n"
        "Total Time (Seconds),Step Time (Seconds),Voltage (mV),Current (mA),Cycle,Step,"
        "Charge Capacity (mAh),Discharge Capacity (mAh),Internal Resistance 1 (mOhm)\n"
        "0,0,3400,100,0,0,0,0,10\n"
        "1,1,3500,-50,0,1,100,0,11\n",
        encoding="utf-8",
    )

    detected = bds.detect(raw)
    df = bds.read(raw, cycler="auto", current_sign="preserve")

    assert detected.cycler == "pec"
    assert df["test_time_s"].to_list() == [0.0, 1.0]
    assert df["voltage_v"].to_list() == [3.4, 3.5]
    assert df["current_a"].to_list() == [0.1, -0.05]
    assert df["charge_capacity_ah"].to_list() == [0.0, 0.1]
    assert df["internal_resistance_ohm"].to_list() == [0.01, 0.011]
    assert validate(df).valid


def test_novonix_preamble_runtime_hours_convert(tmp_path):
    raw = tmp_path / "Nova_Formation_138.csv"
    raw.write_text(
        "metadata\n"
        "Novonix HPC data file\n"
        "Novonix\n"
        "Date and Time,Cycle Number,Step Type,Run Time (h),Step Time (h),Current (A),"
        "Potential (V),Capacity (Ah),Temperature (°C),Circuit Temperature (°C),Energy (Wh),"
        "Step Number,Step position,Power(W)\n"
        "2026-01-01 00:00:00,1,Charge,0,0,0.5,3.4,0,25,26,0,1,1,1.7\n"
        "2026-01-01 00:00:03,1,Discharge,0.001,0.001,-0.5,3.5,0.1,25,26,0.35,2,2,-1.75\n",
        encoding="utf-8",
    )

    detected = bds.detect(raw)
    df = bds.read(raw, cycler="auto")

    assert detected.cycler == "novonix"
    assert df["test_time_s"].to_list() == [0.0, 3.6]
    assert df["current_a"].to_list() == [0.5, -0.5]
    assert validate(df).valid


def test_novonix_data_section_after_protocol_preamble_convert(tmp_path):
    raw = tmp_path / "Nova_Formation_100.csv"
    raw.write_text(
        "\n".join(
            [
                "[Summary]",
                "Novonix HPC data file",
                "Novonix",
                "Protocol: Nova_Formation_100.pro1",
                "[Charge]",
                "   [StepLimits 0 60 0 50 0 9999 0 999 -45 80 -40 60 0 4.8 0 4.6 0 2 0 2]",
                "   [Save data when ΔV > 0.01]",
                "[Data]",
                "Date and Time,Cycle Number,Step Type,Run Time (h),Step Time (h),Current (A),"
                "Potential (V),Capacity (Ah),Temperature (°C),Circuit Temperature (°C),Energy (Wh),"
                "dVdt (I/h),dIdt (V/h),Step Number,Step position",
                "2022-09-13 5:43:03 PM,1,0,0.0,0.0,0.0,0.14838115,0.0,24.08724403,24.937,0.0,0.0,0.0,1,1",
                "2022-09-13 5:43:04 PM,1,0,0.0002917,0.0002917,0.0,0.14838061,0.0,"
                "24.09099388,24.934,0.0,0.0,0.0,1,0",
            ]
        ),
        encoding="utf-8",
    )

    detected = bds.detect(raw)
    df = bds.read(raw, cycler="auto")
    explicit = bds.read(raw, cycler="novonix")

    assert detected.cycler == "novonix"
    assert df["test_time_s"].to_list() == pytest.approx([0.0, 1.05012])
    assert df["step_time_s"].to_list() == pytest.approx([0.0, 1.05012])
    assert df["voltage_v"].to_list() == pytest.approx([0.14838115, 0.14838061])
    assert df["current_a"].to_list() == [0.0, 0.0]
    assert df["cycle_index"].to_list() == [1, 1]
    assert validate(df).valid

    assert explicit["test_time_s"].to_list() == pytest.approx([0.0, 1.05012])
    assert explicit["voltage_v"].to_list() == pytest.approx([0.14838115, 0.14838061])


def test_digatron_metadata_preamble_and_units_row_convert(tmp_path):
    raw = tmp_path / "557_Charge10.csv"
    raw.write_text(
        "\n".join(
            [
                "",
                "Measurement ID,557",
                "Battery Name,LG HG2",
                "Start Time,11/7/2018 10:49:25 AM",
                "Time Stamp,Step,Status,Prog Time,Step Time,Cycle,Cycle Level,Procedure,Voltage,Current,Temperature,Capacity,WhAccu,Cnt,",
                ",,,,,,,,[V],[A],[C],[Ah],[Wh],[Cnt],",
                "11/7/2018 5:08:07 PM,3,PAU,06:18:41.848,00:01:00.001,0,0,NN_Char_Charge,3.38427,0.00000,38.06739,0.00000,0.00000,2.00000,",
                "11/7/2018 5:09:07 PM,3,CHA,06:19:41.847,00:02:00.000,0,0,NN_Char_Charge,3.38478,0.50000,36.49001,0.00100,0.01000,2.00000,",
            ]
        ),
        encoding="utf-8",
    )

    df = bds.read(raw, cycler="generic", current_sign="preserve")

    assert df["test_time_s"].to_list() == pytest.approx([22721.848, 22781.847])
    assert df["voltage_v"].to_list() == pytest.approx([3.38427, 3.38478])
    assert df["current_a"].to_list() == pytest.approx([0.0, 0.5])
    assert validate(df).valid


def test_landt_short_column_export_maps_time_voltage_current(tmp_path):
    raw = tmp_path / "Charge_DR1.5C_Tset20.csv"
    raw.write_text("T1,TA,P,V,I,t\n20,19,1.0,3.5,0.3,0\n20.1,19.1,1.1,3.6,0.3,1\n", encoding="utf-8")

    df = bds.read(raw, cycler="landt", current_sign="preserve")

    assert df["test_time_s"].to_list() == [0.0, 1.0]
    assert df["voltage_v"].to_list() == [3.5, 3.6]
    assert df["current_a"].to_list() == [0.3, 0.3]
    assert validate(df).valid


def test_excel_title_row_promotes_in_sheet_header(tmp_path):
    pd = pytest.importorskip("pandas")
    path = tmp_path / "mmc1.xlsx"
    frame = pd.DataFrame(
        [
            ["under constant 1C current", None, None, None],
            ["sampling time", "current load", "cell1 voltage(V)", "pack volage(V)"],
            ["2016/5/4 14:03:41s", -9.5, 3.4, 10.4],
            ["2016/5/4 14:03:42s", -9.5, 3.5, 10.5],
        ]
    )
    with pd.ExcelWriter(path) as writer:
        frame.to_excel(writer, sheet_name="cc_dischare", index=False, header=False)

    df = bds.read(path, cycler="generic", sheet="cc_dischare", current_sign="preserve")

    assert df["test_time_s"].to_list() == [0.0, 1.0]
    assert df["voltage_v"].to_list() == [10.4, 10.5]
    assert df["current_a"].to_list() == [-9.5, -9.5]
    assert validate(df).valid


def test_eis_excel_zprime_columns_convert(tmp_path):
    pd = pytest.importorskip("pandas")
    path = tmp_path / "CY25_0.25_1.xlsx"
    frame = pd.DataFrame({"Data: Frequency": [10, 1], "Data: Z'": [0.1, 0.2], "Data: Z''": [-0.01, -0.02]})
    with pd.ExcelWriter(path) as writer:
        frame.to_excel(writer, sheet_name="1_1RCPE", index=False)

    df = bds.read_eis(path, sheet="1_1RCPE")

    assert df["Frequency_Hz"].to_list() == [10.0, 1.0]
    assert df["Zre_exp_Ohm"].to_list() == [0.1, 0.2]
    assert bds.validate_eis(df).valid


def test_eis_complex_impedance_uses_sibling_frequency_table(tmp_path):
    impedance = tmp_path / "impedance.csv"
    impedance.write_text(
        "MEASURE_ID,SOC,BATTERY_ID,FREQUENCY_ID,IMPEDANCE_VALUE\n"
        "02_4,100,02,0,(0.11-0.005j)\n"
        "02_4,100,02,1,(0.12-0.006j)\n",
        encoding="utf-8",
    )
    (tmp_path / "frequencies.csv").write_text(
        "FREQUENCY_ID,FREQUENCY_VALUE\n0,0.05\n1,0.1\n", encoding="utf-8"
    )

    df = bds.read_eis(impedance)

    assert df["Frequency_Hz"].to_list() == [0.1, 0.05]
    assert df["Zre_exp_Ohm"].to_list() == [0.12, 0.11]
    assert df["Zim_exp_Ohm"].to_list() == [-0.006, -0.005]
    assert bds.validate_eis(df).valid


def test_basytec_native_current_sign_preserves_charge_positive(tmp_path):
    raw = tmp_path / "basytec_native.csv"
    raw.write_text(
        "run_time,c_vol,c_cur,c_surf_temp,step_type\n00:00:00.000,3.4,0.1,25,1\n00:00:01.000,3.5,-0.2,25,2\n",
        encoding="utf-8",
    )

    df = bds.read(raw, cycler="basytec")

    assert df["current_a"].to_list() == [0.1, -0.2]
    assert validate(df).valid


def test_excel_explicit_sheet_converts_ambiguous_workbook(tmp_path):
    pd = pytest.importorskip("pandas")
    path = tmp_path / "multi.xlsx"
    frame = pd.DataFrame({"Test Time (s)": [0], "Voltage (V)": [3.4], "Current (A)": [0.1]})
    with pd.ExcelWriter(path) as writer:
        frame.to_excel(writer, sheet_name="cell_a", index=False)
        frame.to_excel(writer, sheet_name="cell_b", index=False)

    df = bds.read(path, cycler="generic", sheet="cell_b")

    assert df["voltage_v"].to_list() == [3.4]


def test_cli_convert_eis(tmp_path):
    raw = tmp_path / "eis.csv"
    out = tmp_path / "out.eis.csv"
    raw.write_text("Frequency_Hz,Zreal_Ohm,Zimag_Ohm\n100,0.1,-0.02\n", encoding="utf-8")

    run = subprocess.run(
        [sys.executable, "-m", "battery_data_standard.cli", "convert-eis", str(raw), str(out)],
        text=True,
        capture_output=True,
        check=True,
    )

    assert out.exists()
    assert json.loads(run.stdout)["valid"] is True


def test_matlab_nested_struct_vectors_convert(tmp_path):
    scipy_io = pytest.importorskip("scipy.io")
    raw = tmp_path / "nested.mat"
    scipy_io.savemat(
        raw,
        {
            "Dataset": {
                "time": [10.0, 11.0, 12.0],
                "voltage": [3.4, 3.5, 3.6],
                "current": [-0.1, -0.2, -0.3],
            }
        },
    )

    df = bds.read(raw, cycler="generic", current_sign="preserve", repair_policy="repair")

    assert df["test_time_s"].to_list() == [0.0, 1.0, 2.0]
    assert df["voltage_v"].to_list() == [3.4, 3.5, 3.6]

"""User-facing explain report renderers."""

from __future__ import annotations

import html
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from .exceptions import UnsupportedFeatureError, UnsupportedFormatError
from .export import export_label_for_canonical
from .io import write_json

REPORT_FORMATS = ("json", "html", "xlsx", "pdf")
_BLUE_ACCENT = "#1F4E79"
_BLUE_ACTION = "#1D4ED8"
_BLUE_HEADER = "#EAF2FB"
_BLUE_LINE = "#D8E4F2"
_BLUE_PANEL = "#F7FAFE"


def write_explain_report(report: Any, path: str | Path) -> Path:
    """Write one explain report using the file extension as the format."""
    output_path = Path(path)
    suffix = output_path.suffix.lower().lstrip(".")
    if suffix not in REPORT_FORMATS:
        raise UnsupportedFormatError(
            f"Unsupported explain report format '{suffix}'. Supported formats: {', '.join(REPORT_FORMATS)}."
        )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    payload = _report_dict(report)
    if suffix == "json":
        write_json(output_path, payload)
    elif suffix == "html":
        output_path.write_text(render_explain_html(payload), encoding="utf-8")
    elif suffix == "xlsx":
        _write_explain_xlsx(payload, output_path)
    elif suffix == "pdf":
        _write_explain_pdf(payload, output_path)
    return output_path


def write_conversion_report(report: Any, path: str | Path) -> Path:
    """Write one conversion report using the file extension as the format."""
    output_path = Path(path)
    suffix = output_path.suffix.lower().lstrip(".")
    if suffix not in REPORT_FORMATS:
        raise UnsupportedFormatError(
            f"Unsupported conversion report format '{suffix}'. Supported formats: {', '.join(REPORT_FORMATS)}."
        )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    report_dict = _report_dict(report)
    if suffix == "json":
        write_json(output_path, report_dict)
    else:
        payload = conversion_report_payload(report_dict)
        if suffix == "html":
            output_path.write_text(render_explain_html(payload), encoding="utf-8")
        elif suffix == "xlsx":
            _write_explain_xlsx(payload, output_path)
        elif suffix == "pdf":
            _write_explain_pdf(payload, output_path)
    return output_path


def write_explain_reports(
    report: Any,
    output_dir: str | Path,
    *,
    stem: str | None = None,
    formats: tuple[str, ...] | list[str] = ("json", "html", "xlsx"),
) -> dict[str, str]:
    """Write a set of explain reports and return paths keyed by format."""
    payload = _report_dict(report)
    output_root = Path(output_dir)
    output_root.mkdir(parents=True, exist_ok=True)
    report_stem = _safe_stem(stem or Path(str(payload.get("input_path") or "bds_explain")).stem)
    outputs: dict[str, str] = {}
    for fmt in formats:
        normalized = fmt.lower().lstrip(".")
        path = output_root / f"{report_stem}.report.{normalized}"
        outputs[normalized] = str(write_explain_report(payload, path))
    return outputs


def write_conversion_reports(
    report: Any,
    output_dir: str | Path,
    *,
    stem: str | None = None,
    formats: tuple[str, ...] | list[str] = ("json", "pdf"),
) -> dict[str, str]:
    """Write a set of conversion reports and return paths keyed by format."""
    report_dict = _report_dict(report)
    output_root = Path(output_dir)
    output_root.mkdir(parents=True, exist_ok=True)
    report_stem = _safe_stem(stem or Path(str(report_dict.get("output_path") or "bds_conversion")).stem)
    outputs: dict[str, str] = {}
    for fmt in formats:
        normalized = fmt.lower().lstrip(".")
        path = output_root / f"{report_stem}.report.{normalized}"
        outputs[normalized] = str(write_conversion_report(report_dict, path))
    return outputs


def conversion_report_payload(report: Any) -> dict[str, Any]:
    """Convert a ``ConversionReport`` dictionary into the rendered report payload."""
    payload = _report_dict(report)
    validation = payload.get("validation") or {}
    metadata = payload.get("metadata") or {}
    target = str(metadata.get("export_target") or "bds")
    status = "ok" if validation.get("valid") else "converted-with-issues"
    provenance = payload.get("provenance") or []
    return {
        "report_title": "BDS Conversion Report",
        "input_path": payload.get("input_path"),
        "output_path": payload.get("output_path"),
        "status": status,
        "data_kind": {"kind": "timeseries", "confidence": None, "reason": "conversion completed"},
        "detection": {
            "cycler": payload.get("cycler"),
            "confidence": payload.get("detection_confidence"),
            "reason": "conversion report",
            "candidates": [],
        },
        "selected_adapter": payload.get("cycler"),
        "confidence": payload.get("detection_confidence"),
        "sheet": payload.get("sheet_name"),
        "source_columns": [str(column) for column in metadata.get("raw_columns", [])],
        "canonical_columns": [item.get("column") for item in provenance if item.get("column")],
        "export_columns": list(metadata.get("export_columns") or payload.get("columns") or []),
        "column_mapping": [
            {
                "source": item.get("source"),
                "canonical_column": item.get("column"),
                "export_column": export_label_for_canonical(str(item.get("column") or ""), target=target),
                "source_unit": item.get("source_unit"),
                "transform": item.get("transform"),
            }
            for item in provenance
        ],
        "unit_transforms": [
            item for item in provenance if str(item.get("transform") or "").startswith("unit conversion")
        ],
        "current_sign": payload.get("current_sign"),
        "repair_policy": metadata.get("repair_policy"),
        "validation": validation,
        "warnings": payload.get("warnings") or [],
        "unmapped_columns": payload.get("unmapped_columns") or [],
        "recommended_next_action": "Review validation issues and warnings before downstream analysis."
        if not validation.get("valid")
        else "The converted file is ready for downstream analysis after review of warnings and provenance.",
        "time_sampling": metadata.get("time_sampling"),
    }


def render_explain_html(report: Any) -> str:
    """Render a polished, self-contained HTML explain report."""
    payload = _report_dict(report)
    generated_at = _generated_at()
    title = _report_title(payload)
    status = str(payload.get("status") or "unknown")
    status_class = _status_class(status)
    data_kind = payload.get("data_kind") or {}
    detection = payload.get("detection") or {}
    validation = payload.get("validation") or {}
    summary = _summary_items(payload)
    mapping_rows = payload.get("column_mapping") or []
    candidates = detection.get("candidates") or []
    validation_issues = validation.get("issues") or []
    warnings = payload.get("warnings") or []
    unmapped = payload.get("unmapped_columns") or []
    source_columns = payload.get("source_columns") or []
    output_path = str(payload.get("output_path") or "")
    next_action = str(payload.get("recommended_next_action") or "")
    time_sampling = payload.get("time_sampling")

    summary_cards = "\n".join(
        f"""<div class="metric"><span>{html.escape(label)}</span><strong>{html.escape(value)}</strong></div>"""
        for label, value in summary
    )
    candidate_rows = _html_table_rows(
        candidates,
        ("cycler", "confidence", "reason"),
        empty="No adapter candidates were reported.",
    )
    mapping_table_rows = _html_table_rows(
        mapping_rows,
        ("source", "canonical_column", "export_column", "source_unit", "transform"),
        empty="No column mapping was reported.",
    )
    issue_rows = _html_table_rows(
        validation_issues,
        ("level", "code", "message", "column"),
        empty="No validation issues were reported.",
    )

    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{html.escape(title)}</title>
  <style>
    :root {{
      --ink: #17202a;
      --muted: #5f6b7a;
      --line: {_BLUE_LINE};
      --panel: {_BLUE_PANEL};
      --accent: {_BLUE_ACCENT};
      --ok: {_BLUE_ACTION};
      --warn: #b45309;
      --bad: #b91c1c;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: Arial, Helvetica, sans-serif;
      color: var(--ink);
      background: #ffffff;
      font-size: 14px;
      line-height: 1.45;
    }}
    main {{ max-width: 1180px; margin: 0 auto; padding: 34px 28px 48px; }}
    header {{
      border-bottom: 3px solid var(--accent);
      padding-bottom: 18px;
      margin-bottom: 22px;
    }}
    h1 {{ margin: 0 0 8px; font-size: 30px; line-height: 1.15; }}
    h2 {{ margin: 28px 0 10px; font-size: 18px; color: var(--accent); }}
    h3 {{ margin: 18px 0 8px; font-size: 15px; }}
    p {{ margin: 6px 0; }}
    code {{ font-family: Consolas, Menlo, monospace; font-size: 12px; overflow-wrap: anywhere; }}
    .subtle {{ color: var(--muted); }}
    .badge {{
      display: inline-block;
      border-radius: 999px;
      padding: 4px 10px;
      font-size: 12px;
      font-weight: 700;
      color: #ffffff;
      vertical-align: middle;
    }}
    .badge.ok {{ background: var(--ok); }}
    .badge.warn {{ background: var(--warn); }}
    .badge.bad {{ background: var(--bad); }}
    .badge.neutral {{ background: #64748b; }}
    .result {{
      border: 1px solid var(--line);
      border-left: 5px solid var(--accent);
      background: var(--panel);
      padding: 14px 16px;
      border-radius: 6px;
      margin: 18px 0;
    }}
    .metrics {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
      gap: 10px;
      margin: 18px 0;
    }}
    .metric {{
      border: 1px solid var(--line);
      border-radius: 6px;
      padding: 12px 14px;
      min-height: 74px;
      background: #fff;
    }}
    .metric span {{ display: block; color: var(--muted); font-size: 12px; margin-bottom: 6px; }}
    .metric strong {{ display: block; font-size: 18px; line-height: 1.2; overflow-wrap: anywhere; }}
    .grid-2 {{
      display: grid;
      grid-template-columns: minmax(0, 1fr) minmax(0, 1fr);
      gap: 18px;
    }}
    table {{ width: 100%; border-collapse: collapse; margin: 10px 0 18px; table-layout: fixed; }}
    th, td {{ border-bottom: 1px solid var(--line); padding: 8px 9px; text-align: left; vertical-align: top; }}
    th {{ background: {_BLUE_HEADER}; font-size: 12px; color: #263545; }}
    td {{ overflow-wrap: anywhere; }}
    ul.columns {{
      columns: 3 220px;
      padding-left: 18px;
      margin-top: 8px;
    }}
    li {{ break-inside: avoid; margin-bottom: 3px; }}
    details {{
      border: 1px solid var(--line);
      border-radius: 6px;
      padding: 10px 12px;
      background: #fff;
      margin: 8px 0 14px;
    }}
    summary {{ cursor: pointer; font-weight: 700; color: var(--accent); }}
    @media print {{
      main {{ max-width: none; padding: 20px; }}
      .grid-2 {{ grid-template-columns: 1fr; }}
      details {{ break-inside: avoid; }}
      table {{ page-break-inside: auto; }}
      tr {{ page-break-inside: avoid; }}
    }}
    @media (max-width: 720px) {{
      main {{ padding: 22px 16px 34px; }}
      .grid-2 {{ grid-template-columns: 1fr; }}
      h1 {{ font-size: 24px; }}
    }}
  </style>
</head>
<body>
<main>
  <header>
    <h1>{html.escape(title)} <span class="badge {status_class}">{html.escape(status)}</span></h1>
    <p><strong>Input:</strong> <code>{html.escape(str(payload.get("input_path") or ""))}</code></p>
    {f"<p><strong>Output:</strong> <code>{html.escape(output_path)}</code></p>" if output_path else ""}
    <p class="subtle">Generated: {html.escape(generated_at)}</p>
  </header>

  <section class="result">
    <h2>Report</h2>
    <p>{html.escape(_report_statement(payload))}</p>
    {f"<p><strong>Recommended next action:</strong> <code>{html.escape(next_action)}</code></p>" if next_action else ""}
  </section>

  <section class="metrics">
    {summary_cards}
  </section>

  {_html_time_sampling(time_sampling)}

  <section class="grid-2">
    <div>
      <h2>Data Detection</h2>
      <table>
        <tbody>
          <tr><th>Data kind</th><td>{html.escape(str(data_kind.get("kind") or ""))}</td></tr>
          <tr><th>Kind confidence</th><td>{html.escape(_format_value(data_kind.get("confidence")))}</td></tr>
          <tr><th>Reason</th><td>{html.escape(str(data_kind.get("reason") or ""))}</td></tr>
          <tr><th>Selected adapter</th><td>{html.escape(str(payload.get("selected_adapter") or ""))}</td></tr>
          <tr><th>Adapter confidence</th><td>{html.escape(_format_value(payload.get("confidence")))}</td></tr>
          <tr><th>Sheet</th><td>{html.escape(str(payload.get("sheet") or ""))}</td></tr>
        </tbody>
      </table>
    </div>
    <div>
      <h2>Validation</h2>
      <table>
        <tbody>
          <tr><th>Valid</th><td>{html.escape(_format_value(validation.get("valid")))}</td></tr>
          <tr><th>Schema</th><td>{html.escape(str(validation.get("schema_version") or ""))}</td></tr>
          <tr><th>Rows</th><td>{html.escape(_format_value(validation.get("rows")))}</td></tr>
          <tr><th>Issues</th><td>{len(validation_issues)}</td></tr>
          <tr><th>Warnings</th><td>{len(warnings)}</td></tr>
          <tr><th>Unmapped columns</th><td>{len(unmapped)}</td></tr>
        </tbody>
      </table>
    </div>
  </section>

  <section>
    <h2>Adapter Candidates</h2>
    <table>
      <thead><tr><th>Cycler</th><th>Confidence</th><th>Reason</th></tr></thead>
      <tbody>{candidate_rows}</tbody>
    </table>
  </section>

  <section>
    <h2>Column Mapping</h2>
    <table>
      <thead><tr><th>Source column</th><th>BDS canonical</th><th>Export column</th><th>Source unit</th><th>Transform</th></tr></thead>
      <tbody>{mapping_table_rows}</tbody>
    </table>
  </section>

  <section class="grid-2">
    <div>
      <h2>Warnings</h2>
      {_html_list(warnings, empty="No conversion warnings were reported.")}
    </div>
    <div>
      <h2>Unmapped Columns</h2>
      {_html_list(unmapped, empty="No unmapped columns were reported.")}
    </div>
  </section>

  <section>
    <h2>Validation Issues</h2>
    <table>
      <thead><tr><th>Level</th><th>Code</th><th>Message</th><th>Column</th></tr></thead>
      <tbody>{issue_rows}</tbody>
    </table>
  </section>

  <details>
    <summary>Source Columns ({len(source_columns)})</summary>
    {_html_column_list(source_columns)}
  </details>
</main>
</body>
</html>
"""


def _write_explain_xlsx(payload: dict[str, Any], path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:  # pragma: no cover - dependency is required by package metadata
        raise UnsupportedFeatureError("Excel reports require openpyxl.") from exc

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    title = _report_title(payload)

    fills = {
        "title": PatternFill("solid", fgColor="1F4E79"),
        "header": PatternFill("solid", fgColor="EAF2FB"),
        "ok": PatternFill("solid", fgColor="D9EAF7"),
        "warn": PatternFill("solid", fgColor="FCEBCD"),
        "bad": PatternFill("solid", fgColor="FADADA"),
    }
    title_font = Font(color="FFFFFF", bold=True, size=16)
    header_font = Font(color="263545", bold=True)
    normal = Alignment(vertical="top", wrap_text=True)

    summary["A1"] = title
    summary["A1"].font = title_font
    summary["A1"].fill = fills["title"]
    summary.merge_cells("A1:E1")
    summary["A2"] = "Input"
    summary["B2"] = str(payload.get("input_path") or "")
    summary["A3"] = "Output"
    summary["B3"] = str(payload.get("output_path") or "")
    summary["A4"] = "Generated"
    summary["B4"] = _generated_at()
    summary["A6"] = "Report"
    summary["B6"] = _report_statement(payload)
    summary["A7"] = "Recommended next action"
    summary["B7"] = str(payload.get("recommended_next_action") or "")
    for cell in ("A2", "A3", "A4", "A6", "A7"):
        summary[cell].font = header_font
        summary[cell].fill = fills["header"]
    for row in summary.iter_rows(min_row=1, max_row=7, max_col=5):
        for cell in row:
            cell.alignment = normal

    rows = [["Metric", "Value"], *_summary_items(payload)]
    _write_sheet_table(
        summary, "A9", rows, "SummaryTable", header_fill=fills["header"], header_font=header_font
    )
    time_sampling_rows = _time_sampling_rows(payload.get("time_sampling"))
    if time_sampling_rows:
        _write_named_table(
            wb,
            "Time Sampling",
            ["Field", "Value"],
            time_sampling_rows,
            "TimeSamplingTable",
            header_fill=fills["header"],
            header_font=header_font,
        )

    _write_named_table(
        wb,
        "Adapter Candidates",
        ["Cycler", "Confidence", "Reason"],
        [
            [item.get("cycler"), item.get("confidence"), item.get("reason")]
            for item in (payload.get("detection") or {}).get("candidates", [])
        ],
        "CandidatesTable",
        header_fill=fills["header"],
        header_font=header_font,
    )
    _write_named_table(
        wb,
        "Column Mapping",
        ["Source Column", "BDS Canonical", "Export Column", "Source Unit", "Transform"],
        [
            [
                item.get("source"),
                item.get("canonical_column"),
                item.get("export_column"),
                item.get("source_unit"),
                item.get("transform"),
            ]
            for item in payload.get("column_mapping", [])
        ],
        "ColumnMappingTable",
        header_fill=fills["header"],
        header_font=header_font,
    )
    _write_named_table(
        wb,
        "Validation Issues",
        ["Level", "Code", "Message", "Column"],
        [
            [item.get("level"), item.get("code"), item.get("message"), item.get("column")]
            for item in (payload.get("validation") or {}).get("issues", [])
        ],
        "ValidationIssuesTable",
        header_fill=fills["header"],
        header_font=header_font,
    )
    _write_named_table(
        wb,
        "Warnings",
        ["Warning"],
        [[warning] for warning in payload.get("warnings", [])],
        "WarningsTable",
        header_fill=fills["header"],
        header_font=header_font,
    )
    _write_named_table(
        wb,
        "Unmapped Columns",
        ["Column"],
        [[column] for column in payload.get("unmapped_columns", [])],
        "UnmappedColumnsTable",
        header_fill=fills["header"],
        header_font=header_font,
    )
    _write_named_table(
        wb,
        "Source Columns",
        ["Column"],
        [[column] for column in payload.get("source_columns", [])],
        "SourceColumnsTable",
        header_fill=fills["header"],
        header_font=header_font,
    )

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = normal
        _fit_columns(ws)
        for table in ws.tables.values():
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )

    # Keep the imports above visibly used for static analysis in older openpyxl builds.
    _ = Table
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _write_explain_pdf(payload: dict[str, Any], path: Path) -> None:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError as exc:
        raise UnsupportedFeatureError("PDF reports require the reportlab dependency.") from exc

    styles = getSampleStyleSheet()
    title = _report_title(payload)
    styles["Title"].textColor = colors.HexColor(_BLUE_ACCENT)
    styles["Heading2"].textColor = colors.HexColor(_BLUE_ACCENT)
    cell_style = ParagraphStyle("BDSCell", parent=styles["BodyText"], fontSize=8, leading=10)
    header_style = ParagraphStyle(
        "BDSHeaderCell",
        parent=cell_style,
        fontName="Helvetica-Bold",
        textColor=colors.HexColor("#263545"),
    )
    doc = SimpleDocTemplate(
        str(path),
        pagesize=letter,
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=36,
        title=title,
    )
    story: list[Any] = []
    story.append(Paragraph(title, styles["Title"]))
    story.append(Paragraph(f"Input: {_pdf_escape(str(payload.get('input_path') or ''))}", styles["BodyText"]))
    output_path = str(payload.get("output_path") or "")
    if output_path:
        story.append(Paragraph(f"Output: {_pdf_escape(output_path)}", styles["BodyText"]))
    story.append(Paragraph(f"Generated: {_generated_at()}", styles["BodyText"]))
    story.append(Spacer(1, 12))
    story.append(Paragraph("Report", styles["Heading2"]))
    story.append(Paragraph(_pdf_escape(_report_statement(payload)), styles["BodyText"]))
    next_action = str(payload.get("recommended_next_action") or "")
    if next_action:
        story.append(Paragraph(f"Recommended next action: {_pdf_escape(next_action)}", styles["BodyText"]))
    story.append(Spacer(1, 10))

    story.append(Paragraph("Summary", styles["Heading2"]))
    story.append(
        _pdf_table(
            [["Metric", "Value"], *_summary_items(payload)],
            Table,
            TableStyle,
            colors,
            Paragraph,
            cell_style,
            header_style,
            col_widths=[170, 320],
        )
    )
    story.append(Spacer(1, 10))

    time_sampling_rows = _time_sampling_rows(payload.get("time_sampling"))
    if time_sampling_rows:
        story.append(Paragraph("Time Sampling", styles["Heading2"]))
        story.append(
            _pdf_table(
                [["Field", "Value"], *time_sampling_rows],
                Table,
                TableStyle,
                colors,
                Paragraph,
                cell_style,
                header_style,
                col_widths=[170, 320],
            )
        )
        story.append(Spacer(1, 10))

    story.append(Paragraph("Column Mapping", styles["Heading2"]))
    mapping_rows = [
        ["Source", "BDS Canonical", "Export", "Unit", "Transform"],
        *[
            [
                item.get("source"),
                item.get("canonical_column"),
                item.get("export_column"),
                item.get("source_unit"),
                item.get("transform"),
            ]
            for item in payload.get("column_mapping", [])[:28]
        ],
    ]
    story.append(
        _pdf_table(
            mapping_rows,
            Table,
            TableStyle,
            colors,
            Paragraph,
            cell_style,
            header_style,
            col_widths=[115, 105, 115, 50, 105],
        )
    )

    issues = (payload.get("validation") or {}).get("issues", [])
    if issues:
        story.append(Spacer(1, 10))
        story.append(Paragraph("Validation Issues", styles["Heading2"]))
        issue_rows = [
            ["Level", "Code", "Message", "Column"],
            *[
                [item.get("level"), item.get("code"), item.get("message"), item.get("column")]
                for item in issues[:24]
            ],
        ]
        story.append(
            _pdf_table(
                issue_rows,
                Table,
                TableStyle,
                colors,
                Paragraph,
                cell_style,
                header_style,
                col_widths=[55, 105, 275, 55],
            )
        )

    warnings = payload.get("warnings") or []
    unmapped = payload.get("unmapped_columns") or []
    if warnings or unmapped:
        story.append(Spacer(1, 10))
        story.append(Paragraph("Warnings and Unmapped Columns", styles["Heading2"]))
        story.append(
            Paragraph(
                _pdf_escape("Warnings: " + ("; ".join(warnings) if warnings else "none")), styles["BodyText"]
            )
        )
        story.append(
            Paragraph(
                _pdf_escape("Unmapped columns: " + (", ".join(unmapped) if unmapped else "none")),
                styles["BodyText"],
            )
        )

    path.parent.mkdir(parents=True, exist_ok=True)
    doc.build(story)


def _write_named_table(
    wb: Any,
    sheet_name: str,
    headers: list[str],
    rows: list[list[Any]],
    table_name: str,
    *,
    header_fill: Any,
    header_font: Any,
) -> None:
    ws = wb.create_sheet(sheet_name[:31])
    data = [headers, *(rows or [["" for _ in headers]])]
    _write_sheet_table(ws, "A1", data, table_name, header_fill=header_fill, header_font=header_font)


def _write_sheet_table(
    ws: Any,
    anchor: str,
    rows: list[list[Any]],
    table_name: str,
    *,
    header_fill: Any,
    header_font: Any,
) -> None:
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table

    start_col = ws[anchor].column
    start_row = ws[anchor].row
    for row_idx, row in enumerate(rows, start=start_row):
        for col_idx, value in enumerate(row, start=start_col):
            ws.cell(row=row_idx, column=col_idx, value=_excel_value(value))
    end_row = start_row + len(rows) - 1
    end_col = start_col + len(rows[0]) - 1
    for cell in ws[start_row]:
        if start_col <= cell.column <= end_col:
            cell.font = header_font
            cell.fill = header_fill
    ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    ws.add_table(Table(displayName=table_name, ref=ref))


def _fit_columns(ws: Any) -> None:
    from openpyxl.utils import get_column_letter

    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            max_len = max(max_len, len(str(value)) if value is not None else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 52)
    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 24


def _pdf_table(
    rows: list[list[Any]],
    table_cls: Any,
    style_cls: Any,
    colors: Any,
    paragraph_cls: Any,
    cell_style: Any,
    header_style: Any,
    *,
    col_widths: list[int],
) -> Any:
    clean_rows = [
        [
            paragraph_cls(_pdf_escape(_format_value(cell)), header_style if row_idx == 0 else cell_style)
            for cell in row
        ]
        for row_idx, row in enumerate(rows)
    ]
    table = table_cls(clean_rows, colWidths=col_widths, repeatRows=1)
    table.setStyle(
        style_cls(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(_BLUE_HEADER)),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#263545")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor(_BLUE_LINE)),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(_BLUE_PANEL)]),
            ]
        )
    )
    return table


def _report_title(payload: dict[str, Any]) -> str:
    return str(payload.get("report_title") or "BDS Diagnostic Report")


def _html_time_sampling(value: Any) -> str:
    rows = _time_sampling_rows(value)
    if not rows:
        return ""
    body = "".join(
        f"<tr><th>{html.escape(str(label))}</th><td>{html.escape(_format_value(item))}</td></tr>"
        for label, item in rows
    )
    return f"""
  <section>
    <h2>Time Sampling</h2>
    <table>
      <tbody>{body}</tbody>
    </table>
  </section>
"""


def _time_sampling_rows(value: Any) -> list[list[Any]]:
    if not isinstance(value, dict):
        return []
    rows: list[list[Any]] = [
        ["Status", value.get("status")],
        ["Policy", value.get("policy")],
        ["Expected interval (s)", value.get("expected_interval_s")],
        ["Interval confidence", value.get("interval_confidence")],
        ["Missing sample points", value.get("missing_points")],
        ["Inserted rows", value.get("inserted_rows")],
        ["Interpolation", value.get("interpolation_method")],
    ]
    gaps = value.get("gaps") or []
    if gaps:
        rows.append(["Gaps", gaps])
    return [[label, item] for label, item in rows if item not in (None, "", [])]


def _summary_items(payload: dict[str, Any]) -> list[tuple[str, str]]:
    data_kind = payload.get("data_kind") or {}
    validation = payload.get("validation") or {}
    validation_issues = validation.get("issues") or []
    validation_errors = sum(1 for issue in validation_issues if issue.get("level") == "error")
    validation_warnings = sum(
        1
        for issue in validation_issues
        if issue.get("level") != "error" and issue.get("code") != "missing-optional-column"
    )
    optional_missing = sum(1 for issue in validation_issues if issue.get("code") == "missing-optional-column")
    time_sampling = payload.get("time_sampling") or {}
    return [
        ("Status", str(payload.get("status") or "")),
        ("Data kind", str(data_kind.get("kind") or "")),
        ("Selected adapter", str(payload.get("selected_adapter") or "")),
        ("Confidence", _format_value(payload.get("confidence"))),
        ("Sheet", str(payload.get("sheet") or "")),
        ("Rows", _format_value(validation.get("rows"))),
        ("Valid", _format_value(validation.get("valid"))),
        ("Warnings", str(len(payload.get("warnings") or []))),
        ("Validation errors", str(validation_errors)),
        ("Validation warnings", str(validation_warnings)),
        ("Optional missing", str(optional_missing)),
        ("Unmapped columns", str(len(payload.get("unmapped_columns") or []))),
        ("Current sign", str(payload.get("current_sign") or "")),
        ("Repair policy", str(payload.get("repair_policy") or "")),
        ("Time sampling", str(time_sampling.get("status") or "")),
    ]


def _report_statement(payload: dict[str, Any]) -> str:
    status = str(payload.get("status") or "unknown")
    if status == "ok":
        adapter = payload.get("selected_adapter") or "selected"
        statement = (
            f"The source file was identified as a time-series export and normalized with the {adapter} "
            "adapter. Validation completed successfully; warnings and unmapped columns should be reviewed "
            "before downstream use."
        )
        time_sampling = payload.get("time_sampling") or {}
        if time_sampling.get("status") == "repaired":
            statement += (
                " Missing sample points on the regular time grid were interpolated and recorded in the "
                "Time Sampling section."
            )
        return statement
    if status == "converted-with-issues":
        return (
            "A normalized table was produced, but validation issues remain. The listed issues should be "
            "resolved or documented before the output is used for analysis."
        )
    if status == "eis":
        return "The source file was identified as an EIS table. The EIS conversion path is recommended for standardized impedance output."
    if status == "eis-with-issues":
        return "The source file was identified as EIS data, but the standardized impedance table has validation issues."
    if status == "unsupported":
        return "No conversion was attempted because the file appears to be helper content or an unsupported data family."
    if status == "error":
        return "The diagnostic conversion did not complete. The error message and adapter candidates should be reviewed."
    return "The diagnostic report was generated. Detection, validation, and column mapping sections should be reviewed."


def _status_class(status: str) -> str:
    if status in {"ok", "eis"}:
        return "ok"
    if status in {"converted-with-issues", "eis-with-issues"}:
        return "warn"
    if status == "error":
        return "bad"
    return "neutral"


def _html_table_rows(rows: list[dict[str, Any]], keys: tuple[str, ...], *, empty: str) -> str:
    if not rows:
        return f'<tr><td colspan="{len(keys)}" class="subtle">{html.escape(empty)}</td></tr>'
    body = []
    for row in rows:
        cells = "".join(f"<td>{html.escape(_format_value(row.get(key)))}</td>" for key in keys)
        body.append(f"<tr>{cells}</tr>")
    return "\n".join(body)


def _html_list(items: list[Any], *, empty: str) -> str:
    if not items:
        return f'<p class="subtle">{html.escape(empty)}</p>'
    body = "".join(f"<li>{html.escape(_format_value(item))}</li>" for item in items)
    return f"<ul>{body}</ul>"


def _html_column_list(items: list[Any]) -> str:
    if not items:
        return '<p class="subtle">No source columns were reported.</p>'
    body = "".join(f"<li><code>{html.escape(_format_value(item))}</code></li>" for item in items)
    return f'<ul class="columns">{body}</ul>'


def _report_dict(report: Any) -> dict[str, Any]:
    if isinstance(report, dict):
        return _clean(report)
    if hasattr(report, "to_dict"):
        return _clean(report.to_dict())
    raise TypeError("report must be a dict or an object with to_dict().")


def _clean(value: Any) -> Any:
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, list):
        return [_clean(item) for item in value]
    if isinstance(value, tuple):
        return [_clean(item) for item in value]
    if isinstance(value, dict):
        return {str(key): _clean(item) for key, item in value.items()}
    return value


def _format_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.3f}".rstrip("0").rstrip(".")
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return str(value)


def _excel_value(value: Any) -> str | int | float | bool | None:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=False, sort_keys=True)


def _pdf_escape(value: str) -> str:
    return html.escape(value).replace("\n", "<br/>")


def _safe_stem(value: str) -> str:
    stem = re.sub(r"[^A-Za-z0-9_.-]+", "_", value).strip("._")
    return stem or "bds_explain"


def _generated_at() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")

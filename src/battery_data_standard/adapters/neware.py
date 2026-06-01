"""NEWARE adapter, including flat exports and multi-section files."""

from __future__ import annotations

import csv
import re
from contextlib import suppress
from dataclasses import dataclass
from datetime import datetime
from itertools import pairwise
from pathlib import Path
from typing import Any

import polars as pl

from ..exceptions import UnsupportedFormatError
from ..io import read_table, read_table_with_metadata, read_text, xlsx_sheet_names
from ..reports import ColumnProvenance, DetectionResult
from .base import AdapterResult
from .generic import GenericAdapter


class NewareAdapter(GenericAdapter):
    id = "neware"
    display_name = "NEWARE"
    adapter_version = "1"
    support_tier = "fixture-backed"
    raw_current_sign = "charge-positive"
    signatures = (
        "neware",
        "cycle id",
        "step id",
        "record id",
        "voltage(v)",
        "current(ma)",
        "capacitance_dchg",
        "dcir(o)",
        "chg.capacity",
        "dchg.capacity",
        "datapoint",
        "total time",
        "step type",
        "channel voltage",
        "channel current",
        "电压(v)",
    )
    column_aliases = {
        "test_time_s": (
            "BDS test_time_s",
            "Total Time(s)",
            "Test Time(s)",
            "TotalTime(s)",
            "TotalTime_S",
            "Absolute Time(h:min:s.ms)",
            "Absolute Time (h:min:s.ms)",
            "Total Time",
            "Test Time",
            "总时间(s)",
            "测试时间(s)",
        ),
        "step_time_s": (
            "Time",
            "Time(h:min:s.ms)",
            "Time(s)",
            "Step Time",
            "State Time(s)",
            "Relative Time(s)",
            "Relative Time",
            "Relative Time(h:min:s.ms)",
            "Relative Time (h:min:s.ms)",
            "StepTime(s)",
            "Step Time(s)",
            "时间(s)",
        ),
        "voltage_v": ("Voltage(V)", "Voltage (V)", "Vol(V)", "Voltage", "Channel Voltage", "电压(V)"),
        "current_a": (
            "Current(mA)",
            "Current (mA)",
            "Cur(mA)",
            "Cur (mA)",
            "Cur(A)",
            "Cur (A)",
            "Cur[A]",
            "Curr(A)",
            "Curr (A)",
            "Current(A)",
            "Current (A)",
            "Current",
            "Channel Current",
            "电流(A)",
            "电流(mA)",
        ),
        "cycle_index": ("Cycle ID", "Cycle", "Cycle Index"),
        "step_index": ("Step ID", "Step", "Step Index"),
        "record_index": ("DataPoint", "Data Point", "Record ID", "Record", "Record Index"),
        "date_time": (
            "Realtime",
            "DateTime",
            "Datetime",
            "Absolute Time",
            "Date",
            "Date(h:min:s.ms)",
            "记录时间",
        ),
        "ambient_temperature_deg_c": (
            "Temperature(C)",
            "Temperature(°C)",
            "Temperature 1 (degC)",
            "温度(°C)",
        ),
        "charge_capacity_ah": (
            "Capacitance_Chg(mAh)",
            "Charge Capacity(mAh)",
            "Charge_Capacity(mAh)",
            "Charge_Capacity (mAh)",
            "Chg.Capacity(mAh)",
            "Chg. Capacity(mAh)",
            "Capacity(mAh)",
            "充电容量(mAh)",
        ),
        "discharge_capacity_ah": (
            "Capacitance_DChg(mAh)",
            "Discharge Capacity(mAh)",
            "Discharge_Capacity(mAh)",
            "Discharge_Capacity (mAh)",
            "DChg.Capacity(mAh)",
            "DChg. Capacity(mAh)",
            "放电容量(mAh)",
        ),
        "charge_energy_wh": ("Engy_Chg(mWh)", "Charge Energy(mWh)", "充电能量(mWh)"),
        "discharge_energy_wh": ("Engy_DChg(mWh)", "Discharge Energy(mWh)", "放电能量(mWh)"),
        "power_w": ("Power(mW)",),
        "internal_resistance_ohm": ("DCIR(O)", "DCIR(Ω)"),
    }

    def sniff(self, path: Path, sample: str) -> DetectionResult:
        lower = sample.lower()
        hits = [token for token in self.signatures if token.lower() in lower]
        if self._looks_like_neware_excel(path, sample):
            confidence = 0.92
            reason = "NEWARE Excel workbook detail or record sheets"
        elif self._looks_multisection(sample):
            confidence = 0.9
            reason = "NEWARE multi-section structure"
        elif hits:
            generic_hits = {"voltage(v)", "current(ma)"}
            specific_count = sum(hit.lower() not in generic_hits for hit in hits)
            generic_count = len(hits) - specific_count
            confidence = 0.05 + min(0.75, specific_count * 0.22 + generic_count * 0.06)
            hit_keys = {hit.lower() for hit in hits}
            if {"voltage(v)", "current(ma)"}.issubset(hit_keys):
                confidence = max(confidence, 0.45)
            reason = f"NEWARE signature hits: {', '.join(hits[:5])}"
        else:
            confidence = 0.0
            reason = "no NEWARE signature"
        return DetectionResult(self.id, confidence, reason)

    def read_raw(self, path: Path, options: dict[str, Any] | None = None) -> pl.DataFrame:
        if path.suffix.lower() not in {".xls", ".xlsx"}:
            text, _encoding = read_text(path)
            if self._looks_multisection(text):
                return self._read_multisection(text)
        excel_result = self._read_neware_excel_with_metadata(path, options)
        if excel_result is not None:
            return excel_result.data
        return read_table(path, options=options)

    def read_raw_with_metadata(self, path: Path, options: dict[str, Any] | None = None) -> AdapterResult:
        if path.suffix.lower() not in {".xls", ".xlsx"}:
            text, encoding = read_text(path)
            if self._looks_multisection(text):
                data = self._read_multisection(text)
                return AdapterResult(
                    data,
                    metadata={
                        "source_format": path.suffix.lower().lstrip(".") or "text",
                        "encoding": encoding,
                        "delimiter": ",",
                        "header_row": 2,
                        "neware_layout": "multi-section",
                        "raw_rows": data.height,
                        "raw_columns": list(data.columns),
                    },
                )
        excel_result = self._read_neware_excel_with_metadata(path, options)
        if excel_result is not None:
            return excel_result
        result = read_table_with_metadata(path, options=options)
        return AdapterResult(result.data, metadata=result.metadata)

    def normalize(
        self,
        raw: pl.DataFrame,
        *,
        profile: dict[str, Any] | None = None,
        strict: bool = True,
        keep_raw: bool = False,
        current_sign: str = "charge-positive",
        repair_policy: str = "warn",
    ) -> AdapterResult:
        result = super().normalize(
            raw,
            profile=profile,
            strict=strict,
            keep_raw=keep_raw,
            current_sign=current_sign,
            repair_policy=repair_policy,
        )
        result = self._derive_detail_capacity_energy(raw, result)
        result = self._append_detail_auxiliary_columns(raw, result)
        result = self._append_record_context_columns(raw, result)
        return self._order_output_columns(result)

    @staticmethod
    def _looks_multisection(text: str) -> bool:
        lines = [line for line in text.splitlines()[:50] if line.strip()]
        if len(lines) < 4:
            return False
        return any(line.startswith(",,") for line in lines[3:]) and any(
            "record" in line.lower() for line in lines[:5]
        )

    @staticmethod
    def _looks_like_neware_excel(path: Path, sample: str) -> bool:
        if path.suffix.lower() not in {".xls", ".xlsx"}:
            return False
        sheet_names = _excel_sheet_names(path)
        if _neware_detail_sheets(sheet_names) or _neware_record_sheets(sheet_names):
            return True
        lower = sample.lower()
        has_neware_sheets = "detail_" in lower and ("statis_" in lower or "cycle_" in lower)
        has_detail_columns = all(token in lower for token in ("record index", "cur(a)", "voltage(v)"))
        has_record_sheet = "\nrecord\n" in f"\n{lower}\n" or "\nrecord\t" in f"\n{lower}"
        has_record_columns = all(
            token in lower for token in ("datapoint", "total time", "current(a)", "voltage(v)")
        )
        return has_neware_sheets or has_detail_columns or (has_record_sheet and has_record_columns)

    @staticmethod
    def _read_neware_excel_with_metadata(
        path: Path,
        options: dict[str, Any] | None,
    ) -> AdapterResult | None:
        if path.suffix.lower() not in {".xls", ".xlsx"}:
            return None
        options = dict(options or {})

        try:
            import pandas as pd

            if path.suffix.lower() == ".xlsx":
                workbook = None
                sheet_names = xlsx_sheet_names(path)
            else:
                workbook = pd.ExcelFile(path)
                sheet_names = [str(name) for name in workbook.sheet_names]
        except Exception:
            return None

        record_paths = _coerce_record_paths(options.get("neware_record_paths"))
        requested_sheet = options.get("sheet")
        if requested_sheet is not None:
            requested = str(requested_sheet)
            if _is_neware_record_sheet(requested):
                return _read_neware_record_workbook(
                    path,
                    sheet_names,
                    [requested],
                    pd,
                    selection="explicit-record-sheet",
                    record_paths=record_paths,
                )
            if not _is_neware_detail_sheet(requested):
                return None
            detail_sheets = [requested]
            selection = "explicit-detail-sheet"
        else:
            detail_sheets = _neware_detail_sheets(sheet_names)
            if not detail_sheets:
                record_sheets = _neware_record_sheets(sheet_names)
                if record_sheets:
                    return _read_neware_record_workbook(
                        path,
                        sheet_names,
                        record_sheets,
                        pd,
                        selection="neware-record-sheet"
                        if len(record_sheets) == 1
                        else "neware-split-record-sheets",
                        record_paths=record_paths,
                    )
            selection = "neware-detail-sheet" if len(detail_sheets) == 1 else "neware-split-detail-sheets"
        if not detail_sheets:
            aux_sheets = _neware_auxiliary_sheets(sheet_names)
            if aux_sheets:
                names = ", ".join(aux_sheets)
                raise UnsupportedFormatError(
                    f"NEWARE Excel file {path} only contains auxiliary sheets ({names}); "
                    "convert the primary workbook that contains Detail sheets."
                )
            return None

        missing = [sheet for sheet in detail_sheets if sheet not in sheet_names]
        if missing:
            raise UnsupportedFormatError(
                f"NEWARE Excel file {path} does not contain requested Detail sheet(s): {missing}."
            )

        detail_sheets = _sort_neware_detail_sheets(detail_sheets)
        if workbook is None:
            workbook = pd.ExcelFile(path)
        raw, metadata, warnings = _read_neware_detail_workbook(path, workbook, detail_sheets, pd)
        metadata.update(
            {
                "source_format": path.suffix.lower().lstrip("."),
                "backend": "pandas",
                "sheet_names": sheet_names,
                "selected_sheets": detail_sheets,
                "sheet_name": detail_sheets[0] if len(detail_sheets) == 1 else None,
                "neware_layout": "excel-detail-sheet"
                if len(detail_sheets) == 1
                else "excel-split-detail-sheets",
                "sheet_selection": selection,
                "raw_rows": raw.height,
                "raw_columns": list(raw.columns),
            }
        )
        return AdapterResult(raw, warnings=warnings, metadata=metadata)

    @staticmethod
    def _derive_detail_capacity_energy(raw: pl.DataFrame, result: AdapterResult) -> AdapterResult:
        status_source = _first_existing(raw.columns, ("Status", "Step Type"))
        if status_source is None:
            return result
        df = result.data
        if df.is_empty():
            return result

        status_values = [_classify_neware_status(value) for value in raw[status_source].to_list()]
        derived: list[str] = []

        capacity_col = _first_existing(
            raw.columns,
            ("CapaCity(Ah)", "Capacity(Ah)", "Capacity (Ah)", "Capacity"),
        )
        if capacity_col is not None:
            values = _float_values(raw[capacity_col])
            if "charge_capacity_ah" not in df.columns:
                df = df.with_columns(
                    pl.Series("charge_capacity_ah", _values_for_status(values, status_values, "charge"))
                )
                result.provenance.append(
                    ColumnProvenance(
                        "charge_capacity_ah",
                        f"{capacity_col}|{status_source}",
                        source_unit="Ah",
                        transform="split NEWARE detail capacity by charge status",
                    )
                )
                derived.append(capacity_col)
            if "discharge_capacity_ah" not in df.columns:
                df = df.with_columns(
                    pl.Series(
                        "discharge_capacity_ah",
                        _values_for_status(values, status_values, "discharge"),
                    )
                )
                result.provenance.append(
                    ColumnProvenance(
                        "discharge_capacity_ah",
                        f"{capacity_col}|{status_source}",
                        source_unit="Ah",
                        transform="split NEWARE detail capacity by discharge status",
                    )
                )
                derived.append(capacity_col)

        energy_col = _first_existing(raw.columns, ("Energy(Wh)", "Energy (Wh)", "Energy"))
        if energy_col is not None:
            values = _float_values(raw[energy_col])
            if "charge_energy_wh" not in df.columns:
                df = df.with_columns(
                    pl.Series("charge_energy_wh", _values_for_status(values, status_values, "charge"))
                )
                result.provenance.append(
                    ColumnProvenance(
                        "charge_energy_wh",
                        f"{energy_col}|{status_source}",
                        source_unit="Wh",
                        transform="split NEWARE detail energy by charge status",
                    )
                )
                derived.append(energy_col)
            if "discharge_energy_wh" not in df.columns:
                df = df.with_columns(
                    pl.Series("discharge_energy_wh", _values_for_status(values, status_values, "discharge"))
                )
                result.provenance.append(
                    ColumnProvenance(
                        "discharge_energy_wh",
                        f"{energy_col}|{status_source}",
                        source_unit="Wh",
                        transform="split NEWARE detail energy by discharge status",
                    )
                )
                derived.append(energy_col)

        if not derived:
            return result
        result.data = df
        mapped = set(result.metadata.get("mapped_columns", []))
        mapped.update(derived)
        result.metadata["mapped_columns"] = sorted(mapped)
        unmapped = [col for col in result.metadata.get("unmapped_columns", []) if col not in set(derived)]
        result.metadata["unmapped_columns"] = unmapped
        return result

    @staticmethod
    def _append_detail_auxiliary_columns(raw: pl.DataFrame, result: AdapterResult) -> AdapterResult:
        df = result.data
        if df.is_empty():
            return result

        appended: list[str] = []
        for col in raw.columns:
            if not _is_standard_auxiliary_column(col) or col in df.columns:
                continue
            values = _float_values(raw[col])
            df = df.with_columns(pl.Series(col, values, dtype=pl.Float64))
            result.provenance.append(
                ColumnProvenance(col, col, transform="merged NEWARE auxiliary detail channel")
            )
            appended.append(col)

            if _is_tu1_temperature_column(col) and "temperature_t1_deg_c" not in df.columns:
                df = df.with_columns(pl.Series("temperature_t1_deg_c", values, dtype=pl.Float64))
                result.provenance.append(
                    ColumnProvenance(
                        "temperature_t1_deg_c",
                        col,
                        source_unit="degC",
                        transform="copied from NEWARE auxiliary TU1 temperature channel",
                    )
                )
                appended.append("temperature_t1_deg_c")

        if not appended:
            return result
        result.data = df
        mapped = set(result.metadata.get("mapped_columns", []))
        mapped.update(appended)
        result.metadata["mapped_columns"] = sorted(mapped)
        unmapped = [col for col in result.metadata.get("unmapped_columns", []) if col not in set(appended)]
        result.metadata["unmapped_columns"] = unmapped
        return result

    @staticmethod
    def _append_record_context_columns(raw: pl.DataFrame, result: AdapterResult) -> AdapterResult:
        df = result.data
        if df.is_empty():
            return result

        context_columns = {
            "Step Type": "NEWARE Step Type",
            "NEWARE Step Type": "NEWARE Step Type",
        }
        appended: list[str] = []
        for source, label in context_columns.items():
            if source not in raw.columns or label in df.columns:
                continue
            df = df.with_columns(raw[source].alias(label))
            result.provenance.append(ColumnProvenance(label, source, transform="preserved NEWARE context"))
            appended.append(label)

        if not appended:
            return result
        result.data = df
        mapped = set(result.metadata.get("mapped_columns", []))
        mapped.update(source for source in context_columns if source in raw.columns)
        result.metadata["mapped_columns"] = sorted(mapped)
        unmapped = [
            col
            for col in result.metadata.get("unmapped_columns", [])
            if col not in set(context_columns) and col not in set(appended)
        ]
        result.metadata["unmapped_columns"] = unmapped
        return result

    @staticmethod
    def _order_output_columns(result: AdapterResult) -> AdapterResult:
        preferred = [
            "test_time_s",
            "date_time",
            "unix_time_s",
            "voltage_v",
            "current_a",
            "cycle_index",
            "step_index",
            "record_index",
            "step_time_s",
            "power_w",
            "ambient_temperature_deg_c",
            "temperature_t1_deg_c",
            "charge_capacity_ah",
            "discharge_capacity_ah",
            "charge_energy_wh",
            "discharge_energy_wh",
            "internal_resistance_ohm",
            "NEWARE Step Type",
        ]
        existing = [column for column in preferred if column in result.data.columns]
        existing.extend(column for column in result.data.columns if column not in set(existing))
        if existing != result.data.columns:
            result.data = result.data.select(existing)
        return result

    @staticmethod
    def _read_multisection(text: str) -> pl.DataFrame:
        lines = [line for line in text.splitlines() if line.strip()]
        cycle_header = _parse_csv_line(lines[0])
        step_header = _parse_csv_line(lines[1])
        record_header = _parse_csv_line(lines[2])

        cycle_col = _clean_cell(cycle_header[0]) if cycle_header else "Cycle ID"
        step_col = _clean_cell(step_header[1]) if len(step_header) > 1 else "Step ID"
        record_columns = [_clean_cell(c) for c in record_header]
        if record_columns:
            record_columns[0] = cycle_col or "Cycle ID"
        if len(record_columns) > 1:
            record_columns[1] = step_col or "Step ID"

        ir_index = None
        for idx, name in enumerate(step_header):
            if _clean_cell(name).lower() in {"dcir(o)", "dcir(ω)", "dcir(ohm)"}:
                ir_index = idx
                break
        if ir_index is not None and "DCIR(O)" not in record_columns:
            record_columns.append("DCIR(O)")

        rows: list[list[str | int | None]] = []
        cycle_number: str | int | None = None
        step_number: str | int | None = None
        ir_value: str | None = None
        for line in lines[3:]:
            cells = _parse_csv_line(line)
            if not cells:
                continue
            if line.startswith(',"') or (cells[0] == "" and len(cells) > 1 and cells[1] != ""):
                step_number = cells[1] if len(cells) > 1 else step_number
                if ir_index is not None and ir_index < len(cells):
                    ir_value = cells[ir_index]
                continue
            if line.startswith(",,") or (len(cells) > 2 and cells[0] == "" and cells[1] == ""):
                record = list(cells)
                if len(record) < len(record_columns):
                    record.extend([None] * (len(record_columns) - len(record)))
                record[0] = cycle_number
                if len(record) > 1:
                    record[1] = step_number
                if ir_index is not None:
                    if len(record) == len(record_columns) - 1:
                        record.append(ir_value)
                    else:
                        record[-1] = ir_value
                rows.append(record[: len(record_columns)])
                continue
            cycle_number = cells[0]

        return pl.DataFrame(rows, schema=record_columns, orient="row")


def _parse_csv_line(line: str) -> list[str]:
    return next(csv.reader([line]))


def _clean_cell(value: str) -> str:
    return str(value).replace("\t", "").replace('"', "").strip()


def _read_neware_detail_workbook(
    path: Path,
    workbook: Any,
    detail_sheets: list[str],
    pd: Any,
) -> tuple[pl.DataFrame, dict[str, Any], list[str]]:
    metadata: dict[str, Any] = {}
    warnings: list[str] = []
    frames = []
    for sheet in detail_sheets:
        frame = pd.read_excel(workbook, sheet_name=sheet).dropna(how="all")
        if not frame.empty:
            frames.append(frame)
    if not frames:
        raise UnsupportedFormatError(f"NEWARE Excel file {path} has no rows in Detail sheets.")

    combined = pd.concat(frames, ignore_index=True, sort=False)
    combined, order_metadata, order_warnings = _check_and_sort_detail_rows(combined, pd)
    warnings.extend(order_warnings)
    metadata.update(order_metadata)

    aux_tables, aux_metadata = _read_neware_auxiliary_tables(path, workbook, pd)
    metadata.update(aux_metadata)
    if aux_tables:
        combined = _merge_auxiliary_tables(combined, aux_tables, pd)

    return _pandas_to_polars_frame(combined, pd), metadata, warnings


def _read_neware_record_workbook(
    path: Path,
    sheet_names: list[str],
    record_sheets: list[str],
    pd: Any,
    *,
    selection: str,
    record_paths: list[Path] | None = None,
) -> AdapterResult:
    frames = []
    record_sheet_refs: list[str] = []
    record_path_refs: list[str] = []
    warnings: list[str] = []

    workbook_sources = (
        _explicit_neware_record_workbooks(path, sheet_names, record_paths)
        if record_paths
        else _neware_record_workbooks(path, sheet_names)
    )
    for workbook_path, names in workbook_sources:
        selected = [sheet for sheet in record_sheets if sheet in names] if workbook_path == path else []
        if not selected:
            selected = _neware_record_sheets(names)
        for sheet in selected:
            frame = _read_excel_sheet_as_pandas(workbook_path, sheet, pd).dropna(how="all")
            if frame.empty:
                continue
            frames.append(frame)
            record_sheet_refs.append(f"{workbook_path.name}:{sheet}")
        if selected:
            record_path_refs.append(str(workbook_path))

    if not frames:
        raise UnsupportedFormatError(f"NEWARE Excel file {path} has no rows in record sheets.")

    combined = pd.concat(frames, ignore_index=True, sort=False)
    combined, order_metadata, order_warnings = _check_and_sort_record_rows(combined, pd)
    warnings.extend(order_warnings)

    metadata: dict[str, Any] = {
        "source_format": path.suffix.lower().lstrip("."),
        "backend": "openpyxl" if path.suffix.lower() == ".xlsx" else "pandas",
        "sheet_names": sheet_names,
        "selected_sheets": record_sheets,
        "sheet_name": record_sheets[0] if len(record_sheets) == 1 else None,
        "record_sheets": record_sheet_refs,
        "record_paths": record_path_refs,
        "record_continuation_paths": [item for item in record_path_refs if item != str(path)],
        "neware_layout": "excel-record-sheet",
        "sheet_selection": selection,
    }
    metadata.update(order_metadata)

    step_frame = _read_optional_neware_context_sheet(path, sheet_names, "step", pd)
    cycle_frame = _read_optional_neware_context_sheet(path, sheet_names, "cycle", pd)
    combined, context_metadata = _append_record_context_from_summaries(
        combined,
        step_frame=step_frame,
        cycle_frame=cycle_frame,
        pd=pd,
    )
    metadata.update(context_metadata)
    metadata.update({"raw_rows": len(combined), "raw_columns": [str(col) for col in combined.columns]})
    return AdapterResult(_pandas_to_polars_frame(combined, pd), warnings=warnings, metadata=metadata)


def _coerce_record_paths(value: Any) -> list[Path] | None:
    if value is None:
        return None
    if isinstance(value, (str, Path)):
        return [Path(value)]
    try:
        return [Path(item) for item in value]
    except TypeError:
        return None


def _explicit_neware_record_workbooks(
    path: Path,
    sheet_names: list[str],
    record_paths: list[Path] | None,
):
    seen: set[Path] = set()
    for candidate in [path, *(record_paths or [])]:
        resolved = Path(candidate)
        if resolved in seen or not resolved.is_file():
            continue
        seen.add(resolved)
        names = sheet_names if resolved == path else _excel_sheet_names(resolved)
        if _neware_record_sheets(names):
            yield resolved, names


def _neware_record_workbooks(path: Path, sheet_names: list[str]):
    yield path, sheet_names
    if not any(_is_neware_context_sheet(name) for name in sheet_names):
        return
    for sibling in sorted(
        path.parent.glob(f"{path.stem}_*{path.suffix}"), key=lambda p: _natural_key(p.name)
    ):
        if sibling == path or not sibling.is_file():
            continue
        if not re.match(rf"^{re.escape(path.stem)}_\d+{re.escape(path.suffix)}$", sibling.name, re.I):
            continue
        names = _excel_sheet_names(sibling)
        if _neware_record_sheets(names):
            yield sibling, names


def _excel_sheet_names(path: Path) -> list[str]:
    if path.suffix.lower() == ".xlsx":
        try:
            return xlsx_sheet_names(path)
        except Exception:
            return []
    try:
        import pandas as pd

        return [str(name) for name in pd.ExcelFile(path).sheet_names]
    except Exception:
        return []


@dataclass(frozen=True)
class NewareRecordFileInfo:
    path: str
    source_format: str
    sheet_names: list[str]
    record_columns: list[str]
    row_count: int
    start_date: str | None
    end_date: str | None
    start_point: int | None
    end_point: int | None
    start_total_time_s: float | None
    end_total_time_s: float | None
    has_record: bool
    has_context: bool
    context_sheets: list[str]
    role: str

    def to_dict(self) -> dict[str, Any]:
        return {
            "path": self.path,
            "source_format": self.source_format,
            "sheet_names": self.sheet_names,
            "record_columns": self.record_columns,
            "row_count": self.row_count,
            "start_date": self.start_date,
            "end_date": self.end_date,
            "start_point": self.start_point,
            "end_point": self.end_point,
            "start_total_time_s": self.start_total_time_s,
            "end_total_time_s": self.end_total_time_s,
            "has_record": self.has_record,
            "has_context": self.has_context,
            "context_sheets": self.context_sheets,
            "role": self.role,
        }


def inspect_neware_record_file(path: str | Path) -> dict[str, Any]:
    """Inspect a NEWARE record-style export without loading the full table into memory."""
    info = _inspect_neware_record_file(Path(path))
    return info.to_dict()


def group_neware_record_files(paths: list[str | Path]) -> list[dict[str, Any]]:
    """Group NEWARE record exports into tests using content ranges, not only filenames."""
    infos = [_inspect_neware_record_file(Path(path)) for path in paths]
    candidates = [info for info in infos if info.has_record and info.row_count > 0]
    candidates.sort(key=_neware_info_sort_key)

    groups: list[dict[str, Any]] = []
    for info in candidates:
        duplicate_group = _find_duplicate_group(groups, info)
        if duplicate_group is not None:
            _add_duplicate_or_replace(duplicate_group, info)
            continue

        continuation_group = _find_continuation_group(groups, info)
        if continuation_group is not None:
            continuation_group["segments"].append(info)
            continue

        groups.append({"segments": [info], "duplicates": []})

    return [_finalize_neware_group(index + 1, group) for index, group in enumerate(groups)]


def _inspect_neware_record_file(path: Path) -> NewareRecordFileInfo:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xls"}:
        return _inspect_neware_record_workbook(path)
    if suffix in {".csv", ".txt", ".tsv"}:
        return _inspect_neware_record_csv(path)
    return NewareRecordFileInfo(
        path=str(path),
        source_format=suffix.lstrip(".") or "unknown",
        sheet_names=[],
        record_columns=[],
        row_count=0,
        start_date=None,
        end_date=None,
        start_point=None,
        end_point=None,
        start_total_time_s=None,
        end_total_time_s=None,
        has_record=False,
        has_context=False,
        context_sheets=[],
        role="unsupported",
    )


def _inspect_neware_record_workbook(path: Path) -> NewareRecordFileInfo:
    sheet_names = _excel_sheet_names(path)
    record_sheets = _neware_record_sheets(sheet_names)
    context_sheets = [name for name in sheet_names if _is_neware_context_sheet(name)]
    if not record_sheets:
        return NewareRecordFileInfo(
            path=str(path),
            source_format=path.suffix.lower().lstrip("."),
            sheet_names=sheet_names,
            record_columns=[],
            row_count=0,
            start_date=None,
            end_date=None,
            start_point=None,
            end_point=None,
            start_total_time_s=None,
            end_total_time_s=None,
            has_record=False,
            has_context=bool(context_sheets),
            context_sheets=context_sheets,
            role="unsupported",
        )

    columns, row_count, first, last = _scan_xlsx_record_bounds(path, record_sheets[0])
    return _info_from_record_bounds(
        path=path,
        source_format=path.suffix.lower().lstrip("."),
        sheet_names=sheet_names,
        columns=columns,
        row_count=row_count,
        first=first,
        last=last,
        has_context=bool(context_sheets),
        context_sheets=context_sheets,
    )


def _inspect_neware_record_csv(path: Path) -> NewareRecordFileInfo:
    columns, row_count, first, last = _scan_csv_record_bounds(path)
    has_record = _looks_like_record_columns(columns)
    return _info_from_record_bounds(
        path=path,
        source_format=path.suffix.lower().lstrip(".") or "text",
        sheet_names=[],
        columns=columns if has_record else [],
        row_count=row_count if has_record else 0,
        first=first if has_record else None,
        last=last if has_record else None,
        has_context=False,
        context_sheets=[],
    )


def _scan_xlsx_record_bounds(
    path: Path, sheet: str
) -> tuple[list[str], int, dict[str, Any] | None, dict[str, Any] | None]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet]
        with suppress(Exception):
            worksheet.reset_dimensions()
        header: list[str] | None = None
        row_count = 0
        first: dict[str, Any] | None = None
        last: dict[str, Any] | None = None
        for row in worksheet.iter_rows(values_only=True):
            if _is_empty_row(row):
                continue
            if header is None:
                header = _dedupe_columns(
                    [_clean_header_cell(value, index) for index, value in enumerate(row)]
                )
                continue
            values = list(row[: len(header)])
            if len(values) < len(header):
                values.extend([None] * (len(header) - len(values)))
            if _looks_like_repeated_header(values, header):
                continue
            record = dict(zip(header, values, strict=True))
            row_count += 1
            if first is None:
                first = record
            last = record
        return header or [], row_count, first, last
    finally:
        workbook.close()


def _scan_csv_record_bounds(
    path: Path,
) -> tuple[list[str], int, dict[str, Any] | None, dict[str, Any] | None]:
    sample = path.read_text(encoding="utf-8-sig", errors="replace")[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    row_count = 0
    first: dict[str, Any] | None = None
    last: dict[str, Any] | None = None
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        reader = csv.DictReader(handle, dialect=dialect)
        columns = [str(column) for column in (reader.fieldnames or [])]
        for row in reader:
            if not any(str(value).strip() for value in row.values() if value is not None):
                continue
            row_count += 1
            if first is None:
                first = dict(row)
            last = dict(row)
    return columns, row_count, first, last


def _info_from_record_bounds(
    *,
    path: Path,
    source_format: str,
    sheet_names: list[str],
    columns: list[str],
    row_count: int,
    first: dict[str, Any] | None,
    last: dict[str, Any] | None,
    has_context: bool,
    context_sheets: list[str],
) -> NewareRecordFileInfo:
    date_col = _first_existing(columns, ("Date", "Absolute Time", "Realtime", "DateTime", "Datetime"))
    point_col = _first_existing(columns, ("DataPoint", "Data Point", "Record Index", "Record ID"))
    total_col = _first_existing(columns, ("Total Time", "Total Time(s)", "TotalTime(s)", "Test Time(s)"))
    has_record = _looks_like_record_columns(columns)
    role = "main" if has_context and has_record else "record-continuation" if has_record else "unsupported"
    return NewareRecordFileInfo(
        path=str(path),
        source_format=source_format,
        sheet_names=sheet_names,
        record_columns=columns,
        row_count=row_count,
        start_date=_clean_optional(first.get(date_col)) if first and date_col else None,
        end_date=_clean_optional(last.get(date_col)) if last and date_col else None,
        start_point=_parse_int(first.get(point_col)) if first and point_col else None,
        end_point=_parse_int(last.get(point_col)) if last and point_col else None,
        start_total_time_s=_duration_to_seconds(first.get(total_col)) if first and total_col else None,
        end_total_time_s=_duration_to_seconds(last.get(total_col)) if last and total_col else None,
        has_record=has_record,
        has_context=has_context,
        context_sheets=context_sheets,
        role=role,
    )


def _looks_like_record_columns(columns: list[str]) -> bool:
    slugs = {_slug(column) for column in columns}
    return (
        any(slug in {"datapoint", "recordindex", "recordid"} for slug in slugs)
        and any(slug in {"date", "absolutetime", "realtime", "datetime"} for slug in slugs)
        and any("voltage" in slug or slug in {"volv"} for slug in slugs)
        and any("current" in slug or slug.startswith("cur") for slug in slugs)
    )


def _neware_info_sort_key(info: NewareRecordFileInfo) -> tuple[datetime, int, str]:
    return (
        _parse_datetime_value(info.start_date) or datetime.max,
        info.start_point if info.start_point is not None else 2**63 - 1,
        info.path.lower(),
    )


def _find_duplicate_group(
    groups: list[dict[str, Any]],
    info: NewareRecordFileInfo,
) -> dict[str, Any] | None:
    for group in groups:
        if any(_neware_infos_duplicate(segment, info) for segment in group["segments"]):
            return group
    return None


def _find_continuation_group(
    groups: list[dict[str, Any]],
    info: NewareRecordFileInfo,
) -> dict[str, Any] | None:
    candidates = [
        group for group in groups if group["segments"] and _neware_info_continues(group["segments"][-1], info)
    ]
    if not candidates:
        return None
    return max(candidates, key=lambda group: _neware_info_score(group["segments"][-1]))


def _add_duplicate_or_replace(group: dict[str, Any], info: NewareRecordFileInfo) -> None:
    for index, segment in enumerate(group["segments"]):
        if not _neware_infos_duplicate(segment, info):
            continue
        if _neware_info_score(info) > _neware_info_score(segment):
            group["duplicates"].append(segment)
            group["segments"][index] = info
        else:
            group["duplicates"].append(info)
        return
    group["duplicates"].append(info)


def _neware_infos_duplicate(left: NewareRecordFileInfo, right: NewareRecordFileInfo) -> bool:
    return (
        left.row_count == right.row_count
        and left.start_point == right.start_point
        and left.end_point == right.end_point
        and left.start_date == right.start_date
        and left.end_date == right.end_date
        and left.start_total_time_s == right.start_total_time_s
        and left.end_total_time_s == right.end_total_time_s
    )


def _neware_info_continues(previous: NewareRecordFileInfo, current: NewareRecordFileInfo) -> bool:
    if previous.end_point is not None and current.start_point is not None:
        if current.start_point != previous.end_point + 1:
            return False
    elif previous.end_total_time_s is not None and current.start_total_time_s is not None:
        if current.start_total_time_s < previous.end_total_time_s:
            return False
    else:
        return False

    previous_end = _parse_datetime_value(previous.end_date)
    current_start = _parse_datetime_value(current.start_date)
    if previous_end is not None and current_start is not None and current_start < previous_end:
        return False

    if previous.record_columns and current.record_columns:
        previous_required = {_slug(column) for column in previous.record_columns[:10]}
        current_required = {_slug(column) for column in current.record_columns[:10]}
        if previous_required != current_required:
            return False
    return True


def _neware_info_score(info: NewareRecordFileInfo) -> int:
    score = 0
    if info.source_format in {"xlsx", "xls"}:
        score += 10
    if info.has_context:
        score += 20
    if info.role == "main":
        score += 5
    return score


def _finalize_neware_group(index: int, group: dict[str, Any]) -> dict[str, Any]:
    segments: list[NewareRecordFileInfo] = sorted(group["segments"], key=_neware_info_sort_key)
    duplicates: list[NewareRecordFileInfo] = sorted(group["duplicates"], key=_neware_info_sort_key)
    primary = _select_primary_neware_segment(segments)
    return {
        "group_id": f"neware-test-{index:03d}",
        "primary_path": primary.path,
        "output_stem": Path(primary.path).stem,
        "record_paths": [segment.path for segment in segments],
        "duplicate_paths": [duplicate.path for duplicate in duplicates],
        "segments": [segment.to_dict() for segment in segments],
        "duplicates": [duplicate.to_dict() for duplicate in duplicates],
        "row_count": sum(segment.row_count for segment in segments),
        "start_date": segments[0].start_date if segments else None,
        "end_date": segments[-1].end_date if segments else None,
        "start_point": segments[0].start_point if segments else None,
        "end_point": segments[-1].end_point if segments else None,
        "has_context": any(segment.has_context for segment in segments),
        "order_checks": _neware_group_order_checks(segments),
    }


def _select_primary_neware_segment(segments: list[NewareRecordFileInfo]) -> NewareRecordFileInfo:
    return max(segments, key=_neware_info_score)


def _neware_group_order_checks(segments: list[NewareRecordFileInfo]) -> list[dict[str, Any]]:
    checks: list[dict[str, Any]] = []
    for previous, current in pairwise(segments):
        previous_end = _parse_datetime_value(previous.end_date)
        current_start = _parse_datetime_value(current.start_date)
        date_order_ok = previous_end is None or current_start is None or current_start >= previous_end
        point_continuity_ok = (
            previous.end_point is not None
            and current.start_point is not None
            and current.start_point == previous.end_point + 1
        )
        total_time_order_ok = (
            previous.end_total_time_s is None
            or current.start_total_time_s is None
            or current.start_total_time_s >= previous.end_total_time_s
        )
        checks.append(
            {
                "previous_path": previous.path,
                "current_path": current.path,
                "date_order_ok": date_order_ok,
                "point_continuity_ok": point_continuity_ok,
                "total_time_order_ok": total_time_order_ok,
            }
        )
    return checks


def _read_optional_neware_context_sheet(
    path: Path,
    sheet_names: list[str],
    target: str,
    pd: Any,
) -> Any | None:
    sheet = next((name for name in sheet_names if str(name).strip().lower() == target), None)
    if sheet is None:
        return None
    frame = _read_excel_sheet_as_pandas(path, sheet, pd).dropna(how="all")
    return None if frame.empty else frame


def _read_excel_sheet_as_pandas(path: Path, sheet: str, pd: Any) -> Any:
    if path.suffix.lower() != ".xlsx":
        return pd.read_excel(path, sheet_name=sheet)

    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet]
        with suppress(Exception):
            worksheet.reset_dimensions()
        rows = worksheet.iter_rows(values_only=True)
        header: list[str] | None = None
        columns: dict[str, list[Any]] = {}
        for row in rows:
            if _is_empty_row(row):
                continue
            header = _dedupe_columns([_clean_header_cell(value, index) for index, value in enumerate(row)])
            columns = {name: [] for name in header}
            break
        if header is None:
            return pd.DataFrame()

        width = len(header)
        for row in rows:
            if _is_empty_row(row):
                continue
            values = list(row[:width])
            if len(values) < width:
                values.extend([None] * (width - len(values)))
            if _looks_like_repeated_header(values, header):
                continue
            for name, value in zip(header, values, strict=True):
                columns[name].append(value)
        return pd.DataFrame(columns)
    finally:
        workbook.close()


def _is_empty_row(row: Any) -> bool:
    return not row or all(value is None or str(value).strip() == "" for value in row)


def _clean_header_cell(value: Any, index: int) -> str:
    text = "" if value is None else str(value).strip()
    return text or f"Column_{index + 1}"


def _dedupe_columns(columns: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    output: list[str] = []
    for column in columns:
        count = seen.get(column, 0)
        seen[column] = count + 1
        output.append(column if count == 0 else f"{column}_{count + 1}")
    return output


def _looks_like_repeated_header(values: list[Any], header: list[str]) -> bool:
    comparable = ["" if value is None else str(value).strip() for value in values]
    return comparable == header


def _check_and_sort_record_rows(frame: Any, pd: Any) -> tuple[Any, dict[str, Any], list[str]]:
    metadata: dict[str, Any] = {"time_order_checked": True}
    warnings: list[str] = []
    output = frame

    datapoint_col = _first_existing(output.columns, ("DataPoint", "Data Point", "Record Index", "Record ID"))
    if datapoint_col is not None:
        records = pd.to_numeric(output[datapoint_col], errors="coerce")
        record_order_ok = _series_non_decreasing(records)
        metadata["record_index_monotonic"] = record_order_ok
        if not record_order_ok:
            output = (
                output.assign(__bds_record_sort=records)
                .sort_values("__bds_record_sort", kind="mergesort")
                .drop(columns=["__bds_record_sort"])
                .reset_index(drop=True)
            )
            warnings.append(f"NEWARE record rows were sorted by {datapoint_col}.")

    date_col = _first_existing(output.columns, ("Date", "Absolute Time", "Realtime", "DateTime", "Datetime"))
    if date_col is not None:
        times = pd.to_datetime(output[date_col], errors="coerce")
        date_order_ok = _datetime_series_non_decreasing(times)
        metadata["absolute_time_monotonic"] = date_order_ok
        if not date_order_ok:
            warnings.append("NEWARE record Date values are not monotonic after record ordering.")

        total_time_col = _first_existing(
            output.columns,
            ("Total Time", "Total Time(s)", "TotalTime(s)", "Test Time", "Test Time(s)"),
        )
        if total_time_col is not None:
            total_times = [_duration_to_seconds(value) for value in output[total_time_col].to_list()]
            valid_total_times = [value for value in total_times if value is not None]
            total_time_strict = bool(valid_total_times and all(b > a for a, b in pairwise(valid_total_times)))
            metadata["total_time_strictly_increasing"] = total_time_strict
            if not total_time_strict:
                valid_dates = times.dropna()
                if len(valid_dates) == len(times) and _datetime_series_non_decreasing(times):
                    derived, adjusted = _strict_elapsed_seconds_from_datetimes(times, pd)
                    output["BDS test_time_s"] = derived
                    metadata["test_time_source"] = date_col
                    metadata["test_time_duplicate_timestamps_adjusted"] = adjusted
                    warnings.append(
                        f"NEWARE {total_time_col} was not strictly increasing; "
                        f"derived BDS test_time_s from {date_col}."
                    )

    return output, metadata, warnings


def _append_record_context_from_summaries(
    frame: Any,
    *,
    step_frame: Any | None,
    cycle_frame: Any | None,
    pd: Any,
) -> tuple[Any, dict[str, Any]]:
    metadata: dict[str, Any] = {}
    output = frame
    if step_frame is not None:
        metadata["step_rows"] = len(step_frame)
        output, joined = _append_step_context(output, step_frame, pd)
        metadata["step_context_joined"] = joined
    if cycle_frame is not None:
        metadata["cycle_rows"] = len(cycle_frame)
        output, joined = _append_cycle_context(output, cycle_frame, pd)
        metadata["cycle_context_joined"] = joined
    return output, metadata


def _append_step_context(frame: Any, step_frame: Any, pd: Any) -> tuple[Any, bool]:
    date_col = _first_existing(frame.columns, ("Date", "Absolute Time", "Realtime", "DateTime", "Datetime"))
    start_col = _first_existing(step_frame.columns, ("Oneset Date", "Onset Date", "Start Date", "Start Time"))
    end_col = _first_existing(step_frame.columns, ("End Date", "End Time"))
    if date_col is None or start_col is None or end_col is None:
        return frame, False

    cycle_col = _first_existing(step_frame.columns, ("Cycle Index", "Cycle", "Cycle ID"))
    step_col = _first_existing(step_frame.columns, ("Step Index", "Step", "Step ID"))
    step_number_col = _first_existing(step_frame.columns, ("Step Number", "Step No.", "StepNo"))
    step_type_col = _first_existing(step_frame.columns, ("Step Type", "Status"))

    steps = step_frame.copy()
    steps["__bds_start"] = pd.to_datetime(steps[start_col], errors="coerce")
    steps["__bds_end"] = pd.to_datetime(steps[end_col], errors="coerce")
    keep = ["__bds_start", "__bds_end"]
    if cycle_col is not None:
        steps["__bds_cycle"] = pd.to_numeric(steps[cycle_col], errors="coerce")
        keep.append("__bds_cycle")
    if step_col is not None:
        steps["__bds_step"] = pd.to_numeric(steps[step_col], errors="coerce")
        keep.append("__bds_step")
    if step_number_col is not None:
        steps["__bds_step_number"] = pd.to_numeric(steps[step_number_col], errors="coerce")
        keep.append("__bds_step_number")
    if step_type_col is not None:
        steps["NEWARE Step Type"] = steps[step_type_col]
        keep.append("NEWARE Step Type")
    steps = steps[keep].dropna(subset=["__bds_start"]).sort_values("__bds_start")
    if steps.empty:
        return frame, False

    records = frame.copy()
    records["__bds_pos"] = range(len(records))
    records["__bds_date"] = pd.to_datetime(records[date_col], errors="coerce")
    valid_records = records[["__bds_pos", "__bds_date"]].dropna(subset=["__bds_date"])
    if valid_records.empty:
        return frame, False
    valid_records = valid_records.sort_values("__bds_date")
    merged = pd.merge_asof(
        valid_records,
        steps,
        left_on="__bds_date",
        right_on="__bds_start",
        direction="backward",
    )
    in_interval = merged["__bds_end"].isna() | (merged["__bds_date"] <= merged["__bds_end"])
    context_cols = [column for column in merged.columns if column not in {"__bds_pos", "__bds_date"}]
    merged.loc[~in_interval, context_cols] = None

    context = merged.set_index("__bds_pos")
    output = records.drop(columns=["__bds_date"])
    if "Cycle Index" not in output.columns and "__bds_cycle" in context.columns:
        output = output.merge(
            context["__bds_cycle"].rename("Cycle Index"),
            left_on="__bds_pos",
            right_index=True,
            how="left",
        )
    if "Step Index" not in output.columns and "__bds_step" in context.columns:
        output = output.merge(
            context["__bds_step"].rename("Step Index"),
            left_on="__bds_pos",
            right_index=True,
            how="left",
        )
    if "Step Number" not in output.columns and "__bds_step_number" in context.columns:
        output = output.merge(
            context["__bds_step_number"].rename("Step Number"),
            left_on="__bds_pos",
            right_index=True,
            how="left",
        )
    for column in ("NEWARE Step Type",):
        if column in output.columns or column not in context.columns:
            continue
        output = output.merge(
            context[column].rename(column),
            left_on="__bds_pos",
            right_index=True,
            how="left",
        )
    return output.drop(columns=["__bds_pos"]), True


def _append_cycle_context(frame: Any, cycle_frame: Any, pd: Any) -> tuple[Any, bool]:
    record_cycle_col = _first_existing(frame.columns, ("Cycle Index", "Cycle", "Cycle ID"))
    cycle_col = _first_existing(cycle_frame.columns, ("Cycle Index", "Cycle", "Cycle ID"))
    if record_cycle_col is None or cycle_col is None:
        return frame, False
    record_cycles = pd.to_numeric(frame[record_cycle_col], errors="coerce").dropna()
    summary_cycles = pd.to_numeric(cycle_frame[cycle_col], errors="coerce").dropna()
    if record_cycles.empty or summary_cycles.empty:
        return frame, False
    return frame, bool(set(record_cycles.astype(int)).issubset(set(summary_cycles.astype(int))))


def _check_and_sort_detail_rows(frame: Any, pd: Any) -> tuple[Any, dict[str, Any], list[str]]:
    metadata: dict[str, Any] = {"time_order_checked": True}
    warnings: list[str] = []
    output = frame

    if "Record Index" in output.columns:
        records = pd.to_numeric(output["Record Index"], errors="coerce")
        record_order_ok = _series_non_decreasing(records)
        metadata["record_index_monotonic"] = record_order_ok
        if not record_order_ok:
            output = (
                output.assign(__bds_record_sort=records)
                .sort_values("__bds_record_sort", kind="mergesort")
                .drop(columns=["__bds_record_sort"])
                .reset_index(drop=True)
            )
            warnings.append("NEWARE Detail rows were sorted by Record Index.")

    if "Absolute Time" in output.columns:
        times = pd.to_datetime(output["Absolute Time"], errors="coerce")
        absolute_time_order_ok = _datetime_series_non_decreasing(times)
        metadata["absolute_time_monotonic"] = absolute_time_order_ok
        if not absolute_time_order_ok:
            warnings.append(
                "NEWARE Detail Absolute Time values are not monotonic after Record Index ordering."
            )

    return output, metadata, warnings


def _read_neware_auxiliary_tables(
    path: Path,
    workbook: Any,
    pd: Any,
) -> tuple[dict[str, list[Any]], dict[str, Any]]:
    tables: dict[str, list[Any]] = {}
    auxiliary_sheets: list[str] = []
    auxiliary_paths: list[str] = []

    for workbook_path, handle in _neware_auxiliary_workbooks(path, workbook, pd):
        used_this_workbook = False
        for sheet in _sort_neware_detail_sheets(
            _neware_auxiliary_sheets([str(s) for s in handle.sheet_names])
        ):
            sheet_frame = pd.read_excel(handle, sheet_name=sheet).dropna(how="all")
            if sheet_frame.empty:
                continue
            record_col = _find_record_id_column(sheet_frame.columns)
            if record_col is None:
                continue
            for value_col in _auxiliary_value_columns(sheet_frame.columns):
                label = _standard_auxiliary_column_name(value_col)
                table = sheet_frame[[record_col, value_col]].rename(
                    columns={record_col: "Record ID", value_col: label}
                )
                tables.setdefault(label, []).append(table)
            auxiliary_sheets.append(f"{workbook_path.name}:{sheet}")
            used_this_workbook = True
        if used_this_workbook:
            auxiliary_paths.append(str(workbook_path))

    metadata = {
        "auxiliary_sheets": auxiliary_sheets,
        "auxiliary_paths": auxiliary_paths,
        "auxiliary_columns": sorted(tables),
    }
    return tables, metadata


def _neware_auxiliary_workbooks(path: Path, workbook: Any, pd: Any):
    yield path, workbook
    for sibling in sorted(
        path.parent.glob(f"{path.stem}__*{path.suffix}"), key=lambda p: _natural_key(p.name)
    ):
        if sibling == path or not sibling.is_file():
            continue
        try:
            yield sibling, pd.ExcelFile(sibling)
        except Exception:
            continue


def _merge_auxiliary_tables(frame: Any, tables: dict[str, list[Any]], pd: Any) -> Any:
    if "Record Index" not in frame.columns:
        return frame

    output = frame.copy()
    output["__bds_record_id"] = pd.to_numeric(output["Record Index"], errors="coerce").astype("Int64")
    for label, parts in tables.items():
        if not parts:
            continue
        aux = pd.concat(parts, ignore_index=True, sort=False)
        aux["__bds_record_id"] = pd.to_numeric(aux["Record ID"], errors="coerce").astype("Int64")
        aux = aux[["__bds_record_id", label]].dropna(subset=["__bds_record_id"])
        aux = aux.drop_duplicates("__bds_record_id", keep="last")
        output = output.merge(aux, on="__bds_record_id", how="left")
    return output.drop(columns=["__bds_record_id"])


def _pandas_to_polars_frame(frame: Any, pd: Any) -> pl.DataFrame:
    try:
        return pl.from_pandas(frame)
    except Exception:
        columns: dict[str, Any] = {}
        for column in frame.columns:
            series = frame[column]
            if pd.api.types.is_integer_dtype(series.dtype) or pd.api.types.is_float_dtype(series.dtype):
                columns[str(column)] = pd.to_numeric(series, errors="coerce").to_list()
            else:
                columns[str(column)] = series.astype("object").where(pd.notna(series), None).to_list()
        return pl.DataFrame(columns, infer_schema_length=10000)


def _is_neware_detail_sheet(name: str) -> bool:
    clean = str(name).strip().lower()
    if _is_neware_detail_voltage_sheet(clean) or _is_neware_detail_temperature_sheet(clean):
        return False
    return clean == "detail" or clean.startswith("detail_") or clean.startswith("detail ")


def _is_neware_record_sheet(name: str) -> bool:
    clean = str(name).strip().lower()
    return clean == "record" or clean.startswith("record_") or clean.startswith("record ")


def _is_neware_context_sheet(name: str) -> bool:
    clean = str(name).strip().lower()
    return clean in {"cycle", "step", "test", "unit"}


def _is_neware_detail_voltage_sheet(name: str) -> bool:
    clean = str(name).strip().lower()
    return clean == "detailvol" or clean.startswith("detailvol_") or clean.startswith("detailvol ")


def _is_neware_detail_temperature_sheet(name: str) -> bool:
    clean = str(name).strip().lower()
    return clean == "detailtemp" or clean.startswith("detailtemp_") or clean.startswith("detailtemp ")


def _neware_detail_sheets(sheet_names: list[str]) -> list[str]:
    return _sort_neware_detail_sheets([name for name in sheet_names if _is_neware_detail_sheet(name)])


def _neware_record_sheets(sheet_names: list[str]) -> list[str]:
    return _sort_neware_detail_sheets([name for name in sheet_names if _is_neware_record_sheet(name)])


def _neware_auxiliary_sheets(sheet_names: list[str]) -> list[str]:
    return [
        name
        for name in sheet_names
        if _is_neware_detail_voltage_sheet(name) or _is_neware_detail_temperature_sheet(name)
    ]


def _sort_neware_detail_sheets(sheet_names: list[str]) -> list[str]:
    return sorted(sheet_names, key=_natural_key)


def _natural_key(value: str) -> list[Any]:
    return [int(part) if part.isdigit() else part.lower() for part in re.split(r"(\d+)", str(value))]


def _find_record_id_column(columns: Any) -> str | None:
    for column in columns:
        if str(column).strip().lower() in {"record id", "record index"}:
            return str(column)
    return None


def _auxiliary_value_columns(columns: Any) -> list[str]:
    ignored = {
        "record id",
        "record index",
        "step name",
        "relative time(h:min:s.ms)",
        "realtime",
        "gap of voltage",
        "gap of temperature",
    }
    values: list[str] = []
    for column in columns:
        text = str(column).strip()
        lower = text.lower()
        if lower in ignored or lower.startswith("gap of "):
            continue
        if "auxiliary channel" in lower or re.search(r"\btu\d+\s+[ut]\(", lower):
            values.append(text)
    return values


def _standard_auxiliary_column_name(column: str) -> str:
    text = re.sub(r"\s+", " ", str(column).strip())
    text = text.replace("(V)", " / V")
    text = text.replace("(oC)", " / degC")
    text = text.replace("(°C)", " / degC")
    text = text.replace("(C)", " / degC")
    return text


def _is_standard_auxiliary_column(column: str) -> bool:
    text = str(column)
    return text.startswith("Auxiliary channel ") and (" / V" in text or " / degC" in text)


def _is_tu1_temperature_column(column: str) -> bool:
    text = str(column).lower()
    return "tu1" in text and " / degc" in text


def _series_non_decreasing(series: Any) -> bool:
    valid = series.dropna()
    if len(valid) <= 1:
        return True
    return bool((valid.diff().dropna() >= 0).all())


def _datetime_series_non_decreasing(series: Any) -> bool:
    valid = series.dropna()
    if len(valid) <= 1:
        return True
    return bool((valid.diff().dropna().dt.total_seconds() >= 0).all())


def _strict_elapsed_seconds_from_datetimes(series: Any, pd: Any) -> tuple[list[float | None], bool]:
    valid = series.dropna()
    if valid.empty:
        return [None for _value in series], False
    start = valid.iloc[0]
    elapsed: list[float | None] = []
    previous = float("-inf")
    adjusted = False
    for value in series:
        if pd.isna(value):
            elapsed.append(None)
            continue
        seconds = (value - start).total_seconds()
        if seconds <= previous:
            seconds = previous + 1e-6
            adjusted = True
        elapsed.append(seconds)
        previous = seconds
    return elapsed, adjusted


def _classify_neware_status(value: Any) -> str | None:
    text = str(value).strip().lower()
    if not text:
        return None
    if "dchg" in text or "discharge" in text or "dch" in text:
        return "discharge"
    if "chg" in text or "charge" in text or text.startswith("c"):
        return "charge"
    if text in {"r", "rest", "idle", "pause", "ocv", "open circuit"} or "rest" in text:
        return "rest"
    return None


def _float_values(series: pl.Series) -> list[float | None]:
    values: list[float | None] = []
    for value in series.to_list():
        if value is None:
            values.append(None)
            continue
        try:
            values.append(float(value))
        except (TypeError, ValueError):
            values.append(None)
    return values


def _values_for_status(
    values: list[float | None],
    statuses: list[str | None],
    target: str,
) -> list[float | None]:
    output: list[float | None] = []
    for value, status in zip(values, statuses, strict=True):
        if status == target:
            output.append(value)
        elif value == 0:
            output.append(0.0)
        else:
            output.append(None)
    return output


def _first_existing(columns: Any, candidates: tuple[str, ...]) -> str | None:
    exact = {str(column).strip().lower(): str(column) for column in columns}
    for candidate in candidates:
        found = exact.get(candidate.strip().lower())
        if found is not None:
            return found
    slugs = {_slug(column): str(column) for column in columns}
    for candidate in candidates:
        found = slugs.get(_slug(candidate))
        if found is not None:
            return found
    return None


def _clean_optional(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _parse_int(value: Any) -> int | None:
    if value is None:
        return None
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def _parse_datetime_value(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    if not text:
        return None
    for candidate in (text, text.replace("/", "-"), text.replace("-", ":", 2)):
        with suppress(ValueError):
            return datetime.fromisoformat(candidate)
    with suppress(ValueError):
        return datetime.strptime(text, "%Y-%m-%d %H:%M:%S")
    with suppress(ValueError):
        return datetime.strptime(text, "%Y-%m-%d %H-%M-%S")
    return None


def _duration_to_seconds(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    if ":" not in text:
        try:
            return float(text)
        except ValueError:
            return None
    parts = text.split(":")
    try:
        values = [float(part) for part in parts]
    except ValueError:
        return None
    if len(values) == 3:
        hours, minutes, seconds = values
        return hours * 3600 + minutes * 60 + seconds
    if len(values) == 4:
        days, hours, minutes, seconds = values
        return days * 86400 + hours * 3600 + minutes * 60 + seconds
    return None


def _slug(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value).lower())
